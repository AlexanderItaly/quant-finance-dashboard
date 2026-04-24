#!/usr/bin/env python3
"""
Dashboard Finanziaria Quantitativa — IS20
Dash web app pronta per Render.com
"""


import warnings; warnings.filterwarnings("ignore")
import io, datetime
import numpy as np
import pandas as pd
import plotly
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

import yfinance as yf
from scipy.optimize import minimize
from scipy import stats
import statsmodels.api as sm
from statsmodels.tsa.stattools import adfuller, kpss, acf as sm_acf
from statsmodels.stats.diagnostic import acorr_ljungbox, het_breuschpagan
from statsmodels.tsa.seasonal import STL
from statsmodels.regression.linear_model import OLS
from statsmodels.tools import add_constant
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.stats.stattools import durbin_watson
from arch import arch_model
from sklearn.linear_model import LinearRegression, Ridge
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

import dash
from dash import dcc, html, Input, Output, State, callback_context, dash_table
import dash_bootstrap_components as dbc

import scipy; import seaborn; import sklearn; import arch as _arch
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


# ── Classi e funzioni ──────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════
# CELLA 2 — Definizione di tutte le classi e funzioni
#            (Esegui PRIMA del pannello widget)
# ═══════════════════════════════════════════════════════════════════════════

# ──────────────────────────────────────────────────────────
# UTILITÀ
# ──────────────────────────────────────────────────────────

def _banner(title, icon="📊", c1="#1F4E78", c2="#2E86AB"):
    pass

def load_prices(tickers, start, end):
    if isinstance(tickers, str):
        tickers = [tickers]
    print(f"  📥 Download: {tickers}")
    raw = yf.download(tickers, start=start, end=end,
                      auto_adjust=True, progress=False)
    prices = raw["Close"] if len(tickers) > 1 else raw[["Close"]]
    if len(tickers) == 1:
        prices.columns = tickers
    # ── Rimuovi colonne completamente vuote (ticker falliti su Yahoo) ─────────
    empty_cols = prices.columns[prices.isna().all()].tolist()
    if empty_cols:
        print(f"  ⚠️  Ticker non scaricati (esclusi): {empty_cols}")
        prices = prices.drop(columns=empty_cols)
    if prices.empty or prices.shape[1] == 0:
        raise ValueError(
            f"Nessun dato scaricato per {tickers}. "
            "Verifica che i ticker siano corretti su Yahoo Finance.")
    # ── Rimuovi colonne con meno di 100 giorni validi ────────────────────────
    sparse = prices.columns[prices.notna().sum() < 100].tolist()
    if sparse:
        print(f"  ⚠️  Ticker con dati insufficienti (<100gg, esclusi): {sparse}")
        prices = prices.drop(columns=sparse)
    prices.dropna(how="all", inplace=True)
    prices.ffill(inplace=True)
    prices.bfill(inplace=True)
    print(f"  ✅ {len(prices)} giorni, {prices.shape[1]} titoli validi.")
    return prices

def compute_returns(prices, log=True):
    if log:
        return np.log(prices / prices.shift(1)).dropna()
    return prices.pct_change().dropna()


# ══════════════════════════════════════════════════════════════════════════════
# [A] FRONTIERA EFFICIENTE DI MARKOWITZ
# ══════════════════════════════════════════════════════════════════════════════

class EfficientFrontier:
    """
    Frontiera Efficiente con:
      - Simulazione Monte Carlo (con vincolo max peso per titolo)
      - Ottimizzazione numerica: Max-Sharpe, Min-Var, Target-Return
      - mu_override per ricalcolo con rendimenti ARIMA forecast
    """

    def __init__(self, returns, risk_free_rate=0.045):
        # ── Rimuovi colonne con NaN (ticker falliti) ─────────────────────────
        clean = returns.dropna(axis=1, how="all")
        bad = [c for c in returns.columns if c not in clean.columns]
        if bad:
            print(f"  ⚠️  EfficientFrontier: colonne escluse (solo NaN): {bad}")
        clean = clean.dropna()   # rimuovi righe con qualsiasi NaN rimasto
        if clean.shape[1] < 2:
            raise ValueError(
                "Frontiera Efficiente richiede almeno 2 asset con dati validi. "
                f"Trovati: {clean.columns.tolist()}")
        self.returns = clean
        self.rf = risk_free_rate
        self.n = clean.shape[1]
        self.tickers = clean.columns.tolist()
        self.mu = clean.mean() * 252
        self.cov = clean.cov() * 252
        self.corr = clean.corr()
        self.sim_results = None

    def simulate(self, n_sim=5000, seed=42, max_weight=1.0):
        np.random.seed(seed)
        rets, vols, sharpes, wlist = [], [], [], []
        for _ in range(n_sim):
            w = np.random.dirichlet(np.ones(self.n))
            w = np.clip(w, 0, max_weight)
            w = w / w.sum() if w.sum() > 0 else np.ones(self.n) / self.n
            r = w @ self.mu
            v = np.sqrt(w @ self.cov @ w)
            sh = (r - self.rf) / v
            rets.append(r); vols.append(v); sharpes.append(sh); wlist.append(w)
        self.sim_results = pd.DataFrame({"Return": rets, "Volatility": vols,
                                         "Sharpe": sharpes, "Weights": wlist})
        return self.sim_results

    def _port_stats(self, w, mu_override=None):
        mu = mu_override if mu_override is not None else self.mu
        r = w @ mu; v = np.sqrt(w @ self.cov @ w); sh = (r - self.rf) / v
        return r, v, sh

    def _constraints(self):
        return [{"type": "eq", "fun": lambda w: np.sum(w) - 1}]

    def _bounds(self, max_weight=1.0):
        return [(0.0, max_weight)] * self.n

    def max_sharpe(self, max_weight=1.0, mu_override=None):
        mu = mu_override if mu_override is not None else self.mu
        def neg_sharpe(w): return -self._port_stats(w, mu)[2]
        res = minimize(neg_sharpe, np.ones(self.n) / self.n,
                       method="SLSQP", bounds=self._bounds(max_weight),
                       constraints=self._constraints())
        r, v, sh = self._port_stats(res.x, mu)
        return {"weights": res.x, "return": r, "volatility": v, "sharpe": sh}

    def min_variance(self, max_weight=1.0):
        def port_var(w): return w @ self.cov @ w
        res = minimize(port_var, np.ones(self.n) / self.n,
                       method="SLSQP", bounds=self._bounds(max_weight),
                       constraints=self._constraints())
        r, v, sh = self._port_stats(res.x)
        return {"weights": res.x, "return": r, "volatility": v, "sharpe": sh}

    def efficient_for_target(self, target_return, max_weight=1.0):
        def port_var(w): return w @ self.cov @ w
        constraints = self._constraints() + [
            {"type": "eq", "fun": lambda w: w @ self.mu - target_return}]
        res = minimize(port_var, np.ones(self.n) / self.n,
                       method="SLSQP", bounds=self._bounds(max_weight),
                       constraints=constraints)
        if res.success:
            r, v, sh = self._port_stats(res.x)
            return {"weights": res.x, "return": r, "volatility": v, "sharpe": sh}
        return None

    def compute_frontier_curve(self, n_points=80, max_weight=1.0):
        r_min = self.min_variance(max_weight)["return"]
        r_max = float(self.mu.max())
        frontier = []
        for t in np.linspace(r_min, r_max, n_points):
            p = self.efficient_for_target(t, max_weight)
            if p: frontier.append((p["return"], p["volatility"]))
        return np.array(frontier)

    def plot(self, figsize=(14, 7), max_weight=1.0, mu_override=None):
        if self.sim_results is None:
            self.simulate(max_weight=max_weight)
        ms = self.max_sharpe(max_weight, mu_override)
        mv = self.min_variance(max_weight)
        fc = self.compute_frontier_curve(max_weight=max_weight)

        fig, axes = plt.subplots(1, 2, figsize=figsize)
        ax = axes[0]
        sc = ax.scatter(self.sim_results["Volatility"], self.sim_results["Return"],
                        c=self.sim_results["Sharpe"], cmap="RdYlGn", alpha=0.4, s=8)
        plt.colorbar(sc, ax=ax, label="Sharpe Ratio")
        if len(fc):
            ax.plot(fc[:, 1], fc[:, 0], "b-", lw=2.5, label="Frontiera Efficiente", zorder=5)
        ax.scatter(ms["volatility"], ms["return"], marker="*", s=300, c="gold",
                   zorder=6, edgecolors="k", label=f"Max Sharpe ({ms['sharpe']:.2f})")
        ax.scatter(mv["volatility"], mv["return"], marker="D", s=150, c="cyan",
                   zorder=6, edgecolors="k", label=f"Min Var ({mv['volatility']*100:.1f}%)")
        for i, t in enumerate(self.tickers):
            ax.scatter(float(np.sqrt(self.cov.iloc[i, i])), float(self.mu.iloc[i]),
                       marker="o", s=80, c=COLORS[i % len(COLORS)], zorder=7, edgecolors="k")
            ax.annotate(t, (float(np.sqrt(self.cov.iloc[i, i])), float(self.mu.iloc[i])),
                        textcoords="offset points", xytext=(5, 5), fontsize=7)
        title = "Frontiera Efficiente" + (" (ARIMA-adjusted)" if mu_override is not None else "")
        ax.set_xlabel("Volatilità Annualizzata"); ax.set_ylabel("Rendimento Atteso")
        ax.set_title(title); ax.legend(fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x*100:.0f}%"))
        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x*100:.0f}%"))

        ax2 = axes[1]
        mask = ms["weights"] > 0.005
        ax2.pie(ms["weights"][mask],
                labels=[self.tickers[i] for i in range(self.n) if mask[i]],
                colors=COLORS[:mask.sum()], autopct="%1.1f%%", startangle=140,
                wedgeprops={"edgecolor": "white", "linewidth": 1.5})
        ax2.set_title(f"Pesi Max Sharpe\nRet:{ms['return']*100:.1f}%  "
                      f"Vol:{ms['volatility']*100:.1f}%  SR:{ms['sharpe']:.2f}")
        plt.tight_layout(); plt.show()
        return ms, mv

    def plot_correlation(self, figsize=(9, 7)):
        fig, ax = plt.subplots(figsize=figsize)
        mask = np.triu(np.ones_like(self.corr, dtype=bool))
        sns.heatmap(self.corr, mask=mask, annot=True, fmt=".2f",
                    cmap="RdYlGn", center=0, vmin=-1, vmax=1,
                    square=True, ax=ax, cbar_kws={"shrink": 0.8})
        ax.set_title("Matrice di Correlazione")
        plt.tight_layout(); plt.show()


# ══════════════════════════════════════════════════════════════════════════════
# [B] ARIMA FORECAST
#     Dal video: "processo ARIMA definisce trend, stagionalità, errore
#                 e poi analizza rendimento futuro e volatilità futura"
# ══════════════════════════════════════════════════════════════════════════════

class ARIMAForecaster:
    """
    Stima ARIMA(p,d,q) su ogni asset dei rendimenti e produce la previsione
    del rendimento atteso futuro (annualizzato). Permette di ricalcolare
    la frontiera efficiente con mu aggiornati.
    """

    def __init__(self, returns, horizon=21):
        self.returns = returns
        self.horizon = horizon
        self.forecasts = {}
        self.forecast_vols = {}

    def fit_and_forecast(self, order=(1, 0, 1)):
        print(f"\n🔮 ARIMA{order} — Forecast su {self.returns.shape[1]} asset, horizon={self.horizon}gg")
        for ticker in self.returns.columns:
            series = self.returns[ticker].dropna()
            # ── Salta ticker senza dati sufficienti ──────────────────────────
            if len(series) < 50:
                print(f"   ⚠️  {ticker}: saltato (meno di 50 osservazioni)")
                continue
            try:
                fitted = ARIMA(series, order=order).fit()
                forecast = fitted.forecast(steps=self.horizon)
                self.forecasts[ticker] = float(forecast.mean()) * 252
                self.forecast_vols[ticker] = float(series.tail(63).std()) * np.sqrt(252)
                print(f"   {ticker:12s}  μ_arima: {self.forecasts[ticker]*100:+.2f}%  "
                      f"σ_arima: {self.forecast_vols[ticker]*100:.2f}%")
            except Exception as e:
                if len(series) > 0 and not np.isnan(series.mean()):
                    print(f"   ⚠️  {ticker}: fallback media storica ({str(e)[:60]})")
                    self.forecasts[ticker] = float(series.mean()) * 252
                    self.forecast_vols[ticker] = float(series.std()) * np.sqrt(252)
                else:
                    print(f"   ❌  {ticker}: nessun dato valido, escluso.")
        if not self.forecasts:
            raise ValueError("ARIMA: nessun asset con dati sufficienti.")
        return pd.Series(self.forecasts)

    def plot_forecast_comparison(self, mu_historical, figsize=(12, 5)):
        mu_h = pd.Series({t: float(mu_historical.loc[t]) * 100
                          for t in self.forecasts if t in mu_historical.index})
        mu_f = pd.Series({t: v * 100 for t, v in self.forecasts.items()})
        x = np.arange(len(mu_h)); w = 0.35
        fig, ax = plt.subplots(figsize=figsize)
        ax.bar(x - w/2, mu_h.values, w, label="Rendimento Storico %", color="#2196F3", alpha=0.8)
        ax.bar(x + w/2, mu_f.values, w, label="ARIMA Forecast %", color="#FF5722", alpha=0.8)
        ax.set_xticks(x); ax.set_xticklabels(mu_h.index, rotation=45, ha="right")
        ax.axhline(0, color="k", lw=0.8, linestyle="--")
        ax.set_ylabel("Rendimento % annualizzato")
        ax.set_title("Rendimenti: Storici vs ARIMA Forecast")
        ax.legend(); plt.tight_layout(); plt.show()


# ══════════════════════════════════════════════════════════════════════════════
# [C] GESTIONE OUTLIER CON VARIABILI DUMMY
#     Metodologia da outliers.txt:
#     "Crei una variabile dummy per un periodo temporale e la inserisci
#      nella regressione come regressore legato a COVID/inflazione/guerra.
#      Il modello attribuisce un valore al regressore dummy e così
#      stabilizzi i residui → distribuzione normale."
# ══════════════════════════════════════════════════════════════════════════════

class OutlierDummyHandler:
    """
    Gestisce outlier tramite dummy variables per eventi contingenti.
    Usa Newey-West (HAC) per standard error robusti.
    Produce diagnostica completa: R², DW, JB, Breusch-Pagan, AIC, BIC.
    """

    # ── Dizionario eventi storici (visibili nel grafico allegato) ────────────
    # Copre il periodo 2016-2026 (10 anni)
    EVENTS = {
        # ── Pre-COVID ────────────────────────────────────────────────────────
        "Brexit":                ("2016-06-23", "2016-10-31"),  # referendum + turbolenza
        "Guerra_USA_Cina":       ("2018-07-06", "2019-12-31"),  # dazi reciproci
        # ── COVID cycle ──────────────────────────────────────────────────────
        "COVID19_Crash":         ("2020-02-01", "2020-06-30"),  # crollo + lockdown
        "Ripresa_Post_COVID":    ("2020-07-01", "2021-06-30"),  # rally mega-cap
        # ── Post-COVID ───────────────────────────────────────────────────────
        "Inflazione_USA":        ("2021-10-01", "2022-12-31"),  # CPI >8%, rialzi Fed
        "Guerra_Russia_Ucraina": ("2022-02-24", "2022-12-31"),  # invasione
        "SVB_Crisis":            ("2023-03-01", "2023-04-30"),  # bank run Silicon Valley
        # ── AI & Trump ───────────────────────────────────────────────────────
        "Rally_AI":              ("2023-05-01", "2024-12-31"),  # ChatGPT→ NVDA rally
        "Dazi_Trump":            ("2025-01-20", "2026-04-19"),  # tariffe import USA
    }

    def __init__(self, returns_series, name="SPY"):
        if isinstance(returns_series, pd.DataFrame):
            returns_series = returns_series.squeeze()
        self.series = returns_series.dropna()
        self.name = name

    def build_dummies(self, events_to_include=None):
        if events_to_include is None:
            events_to_include = list(self.EVENTS.keys())
        dummies = pd.DataFrame(index=self.series.index)
        for ev_name in events_to_include:
            if ev_name not in self.EVENTS:
                continue
            start, end = self.EVENTS[ev_name]
            col = ev_name.replace(" ", "_").replace("-", "").replace("/","")
            dummies[col] = (
                (self.series.index >= pd.Timestamp(start)) &
                (self.series.index <= pd.Timestamp(end))
            ).astype(float)
        return dummies

    def fit_with_dummies(self, events_to_include=None, use_newey_west=True):
        dummies = self.build_dummies(events_to_include)
        idx = self.series.index.intersection(dummies.index)
        y = self.series.loc[idx]
        X = add_constant(dummies.loc[idx])

        model = OLS(y, X)
        result = (model.fit(cov_type="HAC", cov_kwds={"maxlags": 5})
                  if use_newey_west else model.fit())
        residuals = result.resid

        jb_stat, jb_p = stats.jarque_bera(residuals)[:2]
        dw_val = durbin_watson(residuals)
        lb_p = acorr_ljungbox(residuals, lags=[10], return_df=True)["lb_pvalue"].values[0]
        try:
            bp_stat, bp_p, _, _ = het_breuschpagan(residuals, X)
        except Exception:
            bp_stat, bp_p = np.nan, np.nan

        diag = {
            "r2": result.rsquared, "adj_r2": result.rsquared_adj,
            "f_stat": result.fvalue, "f_p": result.f_pvalue,
            "aic": result.aic, "bic": result.bic,
            "durbin_watson": dw_val, "jb_stat": jb_stat, "jb_p": jb_p,
            "lb_p": lb_p, "bp_stat": bp_stat, "bp_p": bp_p,
            "skewness": float(stats.skew(residuals)),
            "kurtosis": float(stats.kurtosis(residuals)),
        }
        return result, residuals, diag

    def plot(self, events_to_include=None, figsize=(15, 11)):
        result, residuals, diag = self.fit_with_dummies(events_to_include)
        dummies = self.build_dummies(events_to_include)

        fig = plt.figure(figsize=figsize)
        gs = gridspec.GridSpec(3, 3, figure=fig, hspace=0.5, wspace=0.35)

        # Serie originale con periodi outlier evidenziati
        ax1 = fig.add_subplot(gs[0, :])
        ax1.plot(self.series, lw=0.8, color="#607D8B", alpha=0.6, label="Serie originale")
        ev_colors = ["#FF9800", "#F44336", "#9C27B0", "#2196F3", "#4CAF50"]
        for i, col in enumerate(dummies.columns):
            mask = dummies[col].reindex(self.series.index, fill_value=0) == 1
            if mask.any():
                ax1.fill_between(self.series.index,
                                 self.series.min(), self.series.max(),
                                 where=mask, alpha=0.25,
                                 color=ev_colors[i % len(ev_colors)],
                                 label=col.replace("_", " "))
        ax1.set_title(f"Serie con Periodi Outlier Evidenziati — {self.name}")
        ax1.legend(fontsize=7, ncol=3)

        # Residui dopo correzione dummy
        ax2 = fig.add_subplot(gs[1, :])
        ax2.plot(residuals, lw=0.8, color="#2196F3", alpha=0.7)
        ax2.axhline(0, color="k", lw=0.7, linestyle="--")
        ax2.fill_between(residuals.index, residuals, alpha=0.15, color="#2196F3")
        ax2.set_title("Residui dopo Correzione Dummy (Newey-West HAC)")
        ax2.set_ylabel("Errore residuo")

        # Istogramma residui
        ax3 = fig.add_subplot(gs[2, 0])
        ax3.hist(residuals, bins=60, density=True, color="#4CAF50",
                 edgecolor="white", alpha=0.7)
        xn = np.linspace(residuals.min(), residuals.max(), 300)
        ax3.plot(xn, stats.norm.pdf(xn, residuals.mean(), residuals.std()),
                 "r-", lw=2, label="Normale teorica")
        ax3.set_title("Distribuzione Residui"); ax3.legend(fontsize=8)

        # QQ-plot
        ax4 = fig.add_subplot(gs[2, 1])
        sm.qqplot(residuals, line="s", ax=ax4, alpha=0.4)
        ax4.set_title("QQ-Plot vs Normale")

        # Tabella diagnostica
        ax5 = fig.add_subplot(gs[2, 2])
        ax5.axis("off")
        rows_tbl = [
            ["R²", f"{diag['r2']:.4f}"],
            ["Adj R²", f"{diag['adj_r2']:.4f}"],
            ["F-stat", f"{diag['f_stat']:.2f} (p={diag['f_p']:.3f})"],
            ["AIC", f"{diag['aic']:.1f}"],
            ["BIC", f"{diag['bic']:.1f}"],
            ["Durbin-Watson", f"{diag['durbin_watson']:.3f} {'✅' if 1.5<diag['durbin_watson']<2.5 else '⚠️'}"],
            ["Jarque-Bera p", f"{diag['jb_p']:.4f} {'✅ Normale' if diag['jb_p']>0.05 else '❌'}"],
            ["Ljung-Box p(10)", f"{diag['lb_p']:.4f} {'✅' if diag['lb_p']>0.05 else '❌ AC'}"],
            ["Breusch-Pagan p", f"{diag['bp_p']:.4f} {'✅' if not np.isnan(diag['bp_p']) and diag['bp_p']>0.05 else '⚠️'}"],
            ["Kurtosis", f"{diag['kurtosis']:.3f}"],
            ["Skewness", f"{diag['skewness']:.3f}"],
        ]
        tbl = ax5.table(cellText=rows_tbl, colLabels=["Test", "Valore"],
                        loc="center", cellLoc="left")
        tbl.auto_set_font_size(False); tbl.set_fontsize(8); tbl.scale(1, 1.3)
        ax5.set_title("Diagnostica Newey-West", fontsize=9)

        plt.suptitle(f"Outlier Dummy Variables — {self.name}", fontsize=13, y=1.01)
        plt.tight_layout(); plt.show()

        print("\n📊 Coefficienti Dummy (Newey-West HAC):")
        print(result.summary().tables[1])
        return result, residuals, diag


# ══════════════════════════════════════════════════════════════════════════════
# [D] STYLE ANALYSIS — NEWEY-WEST HAC
#     Dal video: "Style analysis con filtro Newey-West per standard error
#                 robusti. Test: R², DW, JB, Breusch-Pagan, AIC, BIC"
# ══════════════════════════════════════════════════════════════════════════════

class StyleAnalysis:
    """
    Replica un portafoglio target con Sharpe Style Analysis.
    Newey-West (HAC) per SE robusti ad autocorrelazione e eteroschedasticità.
    Diagnostica completa come nel video.
    """

    def __init__(self, returns_df, target_ticker, universe_tickers):
        valid_uni = [t for t in universe_tickers if t in returns_df.columns]
        if not valid_uni:
            raise ValueError(f"Nessun ticker universo valido per Style Analysis.")
        if target_ticker not in returns_df.columns:
            raise ValueError(f"Target '{target_ticker}' non trovato.")
        removed = [t for t in universe_tickers if t not in valid_uni]
        if removed:
            print(f"  ⚠️  StyleAnalysis: ticker rimossi: {removed}")
        combined = returns_df[[target_ticker] + valid_uni].dropna()
        self.target        = combined[target_ticker].values
        self.target_series = combined[target_ticker]
        self.universe      = combined[valid_uni].values
        self.universe_df   = combined[valid_uni]
        self.target_name   = target_ticker
        self.universe_names = valid_uni
        self.n             = len(valid_uni)

    def fit(self, use_newey_west=True, max_lags=5):
        X = add_constant(self.universe_df)
        y = self.target_series
        model = OLS(y, X)
        result = (model.fit(cov_type="HAC", cov_kwds={"maxlags": max_lags})
                  if use_newey_west else model.fit())

        coefs = result.params.values[1:]
        w = np.maximum(coefs, 0)
        w = w / w.sum() if w.sum() > 0 else np.ones(self.n) / self.n

        y_hat = self.universe @ w
        resid = self.target - y_hat
        te = np.std(resid) * np.sqrt(252) * 100
        r2_te = 1 - np.var(resid) / np.var(self.target)

        dw_val = durbin_watson(result.resid)
        jb_stat, jb_p = stats.jarque_bera(result.resid)[:2]
        lb_p = acorr_ljungbox(result.resid, lags=[10], return_df=True)["lb_pvalue"].values[0]
        try:
            bp_stat, bp_p, _, _ = het_breuschpagan(result.resid, X)
        except Exception:
            bp_stat, bp_p = np.nan, np.nan

        diag = {
            "r2": result.rsquared, "adj_r2": result.rsquared_adj,
            "f_stat": result.fvalue, "f_p": result.f_pvalue,
            "aic": result.aic, "bic": result.bic,
            "durbin_watson": dw_val, "jb_stat": jb_stat, "jb_p": jb_p,
            "lb_p": lb_p, "bp_stat": bp_stat, "bp_p": bp_p,
            "kurtosis": float(stats.kurtosis(result.resid)),
            "skewness": float(stats.skew(result.resid)),
            "tracking_error": te, "r2_te": r2_te,
        }
        return {"weights": w, "tracking_error": te, "r2": r2_te,
                "fitted": y_hat, "residuals": resid,
                "ols_result": result, "diagnostics": diag}

    def plot(self, result, figsize=(15, 11)):
        w = result["weights"]
        fitted = result["fitted"]
        resid = result["residuals"]
        diag = result["diagnostics"]

        fig = plt.figure(figsize=figsize)
        gs = gridspec.GridSpec(3, 3, figure=fig, hspace=0.5, wspace=0.35)

        ax1 = fig.add_subplot(gs[0, :2])
        bars = ax1.bar(self.universe_names, w * 100,
                       color=[COLORS[i % len(COLORS)] for i in range(self.n)],
                       edgecolor="white")
        ax1.set_title(f"Style Analysis — Pesi {self.target_name} (Newey-West HAC)")
        ax1.set_ylabel("Peso %")
        for bar, val in zip(bars, w):
            if val > 0.005:
                ax1.text(bar.get_x() + bar.get_width() / 2,
                         bar.get_height() + 0.3,
                         f"{val*100:.1f}%", ha="center", va="bottom", fontsize=8)

        ax1b = fig.add_subplot(gs[0, 2])
        ax1b.axis("off")
        rows_tbl = [
            ["R²", f"{diag['r2']*100:.2f}%"],
            ["Adj R²", f"{diag['adj_r2']*100:.2f}%"],
            ["F-stat", f"{diag['f_stat']:.2f}"],
            ["F p-val", f"{diag['f_p']:.2e}"],
            ["AIC", f"{diag['aic']:.1f}"],
            ["BIC", f"{diag['bic']:.1f}"],
            ["DW", f"{diag['durbin_watson']:.3f} {'✅' if 1.5<diag['durbin_watson']<2.5 else '⚠️'}"],
            ["JB p", f"{diag['jb_p']:.4f} {'✅' if diag['jb_p']>0.05 else '❌'}"],
            ["LB p(10)", f"{diag['lb_p']:.4f} {'✅' if diag['lb_p']>0.05 else '❌'}"],
            ["BP p", f"{diag['bp_p']:.4f} {'✅' if not np.isnan(diag['bp_p']) and diag['bp_p']>0.05 else '⚠️'}"],
            ["TE Ann.", f"{diag['tracking_error']:.3f}%"],
            ["Kurtosis", f"{diag['kurtosis']:.3f}"],
        ]
        tbl = ax1b.table(cellText=rows_tbl, colLabels=["Test", "Valore"],
                         loc="center", cellLoc="left")
        tbl.auto_set_font_size(False); tbl.set_fontsize(7.5); tbl.scale(1, 1.25)
        ax1b.set_title("Diagnostica Newey-West", fontsize=9)

        ax2 = fig.add_subplot(gs[1, :])
        ax2.plot(np.cumsum(self.target), label=f"{self.target_name} (originale)",
                 lw=2, color="#2196F3")
        ax2.plot(np.cumsum(fitted), label="Style Analysis Replica",
                 lw=2, color="#FF5722", linestyle="--")
        ax2.set_title(f"Performance Cumulativa — TE: {diag['tracking_error']:.2f}%  "
                      f"R²: {diag['r2']*100:.2f}%")
        ax2.legend(); ax2.set_ylabel("Rendimento Log Cumulativo")

        ax3 = fig.add_subplot(gs[2, 0])
        ax3.plot(resid, lw=0.8, color="#9C27B0", alpha=0.7)
        ax3.axhline(0, color="k", lw=0.8, linestyle="--")
        ax3.set_title("Residui (Tracking Error)")

        ax4 = fig.add_subplot(gs[2, 1])
        ax4.hist(resid, bins=50, color="#4CAF50", edgecolor="white", alpha=0.7, density=True)
        xn = np.linspace(resid.min(), resid.max(), 300)
        ax4.plot(xn, stats.norm.pdf(xn, resid.mean(), resid.std()), "r-", lw=2)
        ax4.set_title("Distribuzione Residui")

        ax5 = fig.add_subplot(gs[2, 2])
        sm.qqplot(pd.Series(resid), line="s", ax=ax5, alpha=0.4)
        ax5.set_title("QQ-Plot")

        plt.suptitle(f"Style Analysis: {self.target_name}", fontsize=14, y=1.01)
        plt.tight_layout(); plt.show()


# ══════════════════════════════════════════════════════════════════════════════
# [E] ETF REPLICATOR
# ══════════════════════════════════════════════════════════════════════════════

class ETFReplicator:
    def __init__(self, returns_df, target_ticker, universe_tickers):
        # ── Filtra universe_tickers a quelli effettivamente presenti ─────────
        valid_uni = [t for t in universe_tickers if t in returns_df.columns]
        if not valid_uni:
            raise ValueError(
                f"Nessun ticker dell'universo ETF trovato nel DataFrame. "
                f"Richiesti: {universe_tickers}")
        if target_ticker not in returns_df.columns:
            raise ValueError(f"ETF target '{target_ticker}' non trovato.")
        if valid_uni != list(universe_tickers):
            removed = [t for t in universe_tickers if t not in valid_uni]
            print(f"  ⚠️  ETFReplicator: ticker rimossi (dati mancanti): {removed}")

        # ── Allinea indici e rimuovi righe con NaN ───────────────────────────
        combined = returns_df[[target_ticker] + valid_uni].dropna()
        if len(combined) < 50:
            raise ValueError(
                f"Troppi pochi dati dopo pulizia NaN: {len(combined)} righe. "
                "Allarga il periodo o riduci i ticker dell'universo.")

        self.target        = combined[target_ticker].values
        self.universe      = combined[valid_uni].values
        self.target_name   = target_ticker
        self.universe_names = valid_uni
        self.n             = len(valid_uni)
        print(f"  ✅ ETFReplicator: {len(combined)} osservazioni, "
              f"{self.n} strumenti nell'universo.")

    def _norm(self, w):
        w = np.maximum(w, 0)
        return w / w.sum() if w.sum() > 0 else np.ones(self.n) / self.n

    def _metrics(self, w):
        yh = self.universe @ w
        te = np.std(self.target - yh) * np.sqrt(252) * 100
        r2 = 1 - np.var(self.target - yh) / np.var(self.target)
        return yh, te, r2

    def replicate_ols(self):
        reg = LinearRegression(fit_intercept=False).fit(self.universe, self.target)
        w = self._norm(reg.coef_)
        yh, te, r2 = self._metrics(w)
        return {"weights": w, "tracking_error": te, "r2": r2,
                "fitted": yh, "residuals": self.target - yh}

    def replicate_ridge(self, alpha=0.01):
        reg = Ridge(alpha=alpha, fit_intercept=False).fit(self.universe, self.target)
        w = self._norm(reg.coef_)
        yh, te, r2 = self._metrics(w)
        return {"weights": w, "tracking_error": te, "r2": r2,
                "fitted": yh, "residuals": self.target - yh}

    def replicate_optimize(self):
        def obj(w): return np.std(self.target - self.universe @ w)
        res = minimize(obj, np.ones(self.n) / self.n, method="SLSQP",
                       bounds=[(0, 1)] * self.n,
                       constraints=[{"type": "eq", "fun": lambda w: np.sum(w) - 1}],
                       options={"maxiter": 1000})
        yh, te, r2 = self._metrics(res.x)
        return {"weights": res.x, "tracking_error": te, "r2": r2,
                "fitted": yh, "residuals": self.target - yh}

    def compare_methods(self):
        ols = self.replicate_ols()
        ridge = self.replicate_ridge()
        opt = self.replicate_optimize()
        print("\n📊 Confronto Metodi Replica ETF:")
        print(pd.DataFrame({
            "Metodo": ["OLS", "Ridge", "Min-TE Opt"],
            "Tracking Error %": [ols["tracking_error"], ridge["tracking_error"], opt["tracking_error"]],
            "R²": [ols["r2"], ridge["r2"], opt["r2"]]}).to_string(index=False))
        return ols, ridge, opt

    def plot(self, result, method_name="OLS", figsize=(15, 10)):
        w = result["weights"]; fitted = result["fitted"]; resid = result["residuals"]
        fig = plt.figure(figsize=figsize)
        gs = gridspec.GridSpec(3, 2, figure=fig, hspace=0.45, wspace=0.35)
        ax1 = fig.add_subplot(gs[0, :])
        bars = ax1.bar(self.universe_names, w * 100,
                       color=[COLORS[i % len(COLORS)] for i in range(self.n)],
                       edgecolor="white")
        ax1.set_title(f"Pesi Replica {self.target_name} — {method_name}")
        ax1.set_ylabel("Peso %")
        for bar, val in zip(bars, w):
            if val > 0.005:
                ax1.text(bar.get_x() + bar.get_width() / 2,
                         bar.get_height() + 0.3,
                         f"{val*100:.1f}%", ha="center", va="bottom", fontsize=8)
        ax2 = fig.add_subplot(gs[1, :])
        ax2.plot(np.cumsum(self.target), label=f"{self.target_name}", lw=2, color="#2196F3")
        ax2.plot(np.cumsum(fitted), label="Replica", lw=2, color="#FF5722", linestyle="--")
        ax2.set_title(f"Performance Cumulativa — TE:{result['tracking_error']:.2f}%  R²:{result['r2']:.4f}")
        ax2.legend()
        ax3 = fig.add_subplot(gs[2, 0])
        ax3.plot(resid, lw=0.8, color="#9C27B0", alpha=0.7)
        ax3.axhline(0, color="k", lw=0.8, linestyle="--")
        ax3.set_title("Tracking Error nel Tempo")
        ax4 = fig.add_subplot(gs[2, 1])
        ax4.hist(resid, bins=50, color="#4CAF50", edgecolor="white", alpha=0.7, density=True)
        xn = np.linspace(resid.min(), resid.max(), 300)
        ax4.plot(xn, stats.norm.pdf(xn, resid.mean(), resid.std()), "r-", lw=2)
        ax4.set_title("Distribuzione Residui")
        plt.suptitle(f"Replica ETF: {self.target_name}", fontsize=14)
        plt.tight_layout(); plt.show()


# ══════════════════════════════════════════════════════════════════════════════
# [F] ANALISI STOCASTICA (STL, ADF/KPSS, ACF/PACF, Normalità, White Noise)
# ══════════════════════════════════════════════════════════════════════════════

class StochasticAnalyzer:
    def __init__(self, series, name="Serie"):
        if isinstance(series, pd.DataFrame):
            series = series.squeeze()
        self.series = series.dropna()
        self.name = name

    def decompose(self, period=252, figsize=(14, 9)):
        print(f"\n🔍 Decomposizione STL: {self.name}")
        result = STL(self.series, period=period, robust=True).fit()
        fig, axes = plt.subplots(4, 1, figsize=figsize, sharex=True)
        for ax, (data, title, color) in zip(axes, [
            (self.series, "Serie Originale", "#2196F3"),
            (result.trend, "Trend", "#FF5722"),
            (result.seasonal, "Stagionalità", "#4CAF50"),
            (result.resid, "Residuo", "#9C27B0")]):
            ax.plot(data, lw=1.2, color=color)
            ax.set_title(title, fontsize=10)
            ax.axhline(0, color="k", lw=0.5, linestyle="--")
        fig.suptitle(f"Decomposizione STL — {self.name}", fontsize=13)
        plt.tight_layout(); plt.show()
        tot = result.trend.var() + result.seasonal.var() + result.resid.var()
        print(f"   Trend:{result.trend.var()/tot*100:.1f}%  "
              f"Stagionalità:{result.seasonal.var()/tot*100:.1f}%  "
              f"Residuo:{result.resid.var()/tot*100:.1f}%")
        return result

    def stationarity_tests(self, series=None):
        s = (series if series is not None else self.series).dropna()
        print(f"\n🔍 Stazionarietà: {self.name}")
        adf_stat, adf_p, _, _, adf_cv, _ = adfuller(s, autolag="AIC")
        kpss_stat, kpss_p, _, _ = kpss(s, regression="c", nlags="auto")
        print(f"  ADF  p={adf_p:.4f}  {'✅ Stazionaria' if adf_p<0.05 else '❌'}")
        print(f"  KPSS p={kpss_p:.4f}  {'✅ Stazionaria' if kpss_p>0.05 else '❌'}")
        return {"adf_p": adf_p, "kpss_p": kpss_p}

    def autocorrelation_analysis(self, n_lags=40, figsize=(14, 6)):
        s = self.series.dropna()
        fig, axes = plt.subplots(1, 2, figsize=figsize)
        plot_acf(s, lags=n_lags, ax=axes[0], alpha=0.05, title=f"ACF — {self.name}")
        plot_pacf(s, lags=n_lags // 2, ax=axes[1], alpha=0.05, title=f"PACF — {self.name}")
        plt.tight_layout(); plt.show()
        lb = acorr_ljungbox(s, lags=[5, 10, 20, 30], return_df=True)
        print(f"\n  Ljung-Box p: {lb['lb_pvalue'].round(4).tolist()}")
        return lb

    def normality_test(self, figsize=(13, 5)):
        s = self.series.dropna()
        jb_stat, jb_p = stats.jarque_bera(s)[:2]
        sw_stat, sw_p = stats.shapiro(s[:5000] if len(s) > 5000 else s)
        print(f"\n🔍 Normalità: JB p={jb_p:.4f}  SW p={sw_p:.4f}  "
              f"Kurt={stats.kurtosis(s):.3f}  Skew={stats.skew(s):.3f}")
        fig, axes = plt.subplots(1, 2, figsize=figsize)
        axes[0].hist(s, bins=60, density=True, color="#2196F3", alpha=0.7, edgecolor="white")
        xn = np.linspace(s.min(), s.max(), 300)
        axes[0].plot(xn, stats.norm.pdf(xn, s.mean(), s.std()), "r-", lw=2)
        axes[0].set_title(f"Istogramma — {self.name}")
        sm.qqplot(s, line="s", ax=axes[1], alpha=0.4)
        axes[1].set_title("QQ-Plot vs Normale")
        plt.tight_layout(); plt.show()

    def transform_to_white_noise(self, figsize=(14, 10)):
        s = self.series.dropna()
        fig, axes = plt.subplots(3, 2, figsize=figsize)

        def _plot(ah, aa, data, title):
            ah.hist(data, bins=50, density=True, color="#4CAF50", alpha=0.7, edgecolor="white")
            xn = np.linspace(data.min(), data.max(), 300)
            ah.plot(xn, stats.norm.pdf(xn, data.mean(), data.std()), "r-", lw=2)
            ah.set_title(title, fontsize=9)
            acf_v = sm_acf(data, nlags=20, fft=True)[1:]
            aa.bar(range(1, 21), acf_v, color="#9C27B0", alpha=0.7)
            conf = 1.96 / np.sqrt(len(data))
            aa.axhline(conf, color="r", linestyle="--", lw=1)
            aa.axhline(-conf, color="r", linestyle="--", lw=1)
            aa.set_ylim(-0.3, 0.3); aa.set_title("ACF", fontsize=9)

        _plot(axes[0, 0], axes[0, 1], s.values, "1. Serie Originale")
        d1 = s.diff().dropna()
        _plot(axes[1, 0], axes[1, 1], d1.values, "2. Prima Differenza")
        d1s = (d1 - d1.mean()) / d1.std()
        _plot(axes[2, 0], axes[2, 1], d1s.values, "3. Diff. Standardizzata")
        plt.suptitle(f"Pipeline White Noise — {self.name}", fontsize=12)
        plt.tight_layout(); plt.show()


# ══════════════════════════════════════════════════════════════════════════════
# [G] VOLATILITÀ ROLLING E GARCH
# ══════════════════════════════════════════════════════════════════════════════

class VolatilityAnalyzer:
    def __init__(self, returns, name="Rendimenti", window=21, rf=0.045):
        if isinstance(returns, pd.DataFrame):
            returns = returns.squeeze()
        self.returns = returns.dropna()
        self.name = name; self.window = window; self.rf = rf

    def rolling_volatility(self, figsize=(14, 8)):
        rv = self.returns.rolling(self.window).std() * np.sqrt(252) * 100
        rm = self.returns.rolling(self.window).mean() * 252 * 100
        rs = (rm - self.rf * 100) / rv
        fig, axes = plt.subplots(3, 1, figsize=figsize, sharex=True)
        axes[0].plot(rv, color="#FF5722", lw=1.2)
        axes[0].fill_between(rv.index, rv, alpha=0.2, color="#FF5722")
        axes[0].set_title(f"Volatilità Rolling {self.window}gg — {self.name}"); axes[0].set_ylabel("Vol %")
        axes[1].plot(rm, color="#2196F3", lw=1.2)
        axes[1].axhline(0, color="k", lw=0.8, linestyle="--")
        axes[1].fill_between(rm.index, rm.clip(lower=0), alpha=0.3, color="#4CAF50")
        axes[1].fill_between(rm.index, rm.clip(upper=0), alpha=0.3, color="#F44336")
        axes[1].set_title(f"Rendimento Rolling — {self.name}")
        axes[2].plot(rs, color="#9C27B0", lw=1.2)
        axes[2].axhline(1.0, color="g", lw=1, linestyle="--", label="SR=1")
        axes[2].axhline(0, color="r", lw=0.8, linestyle="--")
        axes[2].set_title("Sharpe Ratio Rolling"); axes[2].legend()
        plt.tight_layout(); plt.show()
        print(f"\n📊 Vol media:{rv.mean():.2f}%  Max:{rv.max():.2f}% ({rv.idxmax().date()})  SR medio:{rs.mean():.2f}")
        return rv, rm, rs

    def garch_model(self, p=1, q=1, figsize=(14, 8)):
        print(f"\n🔍 GARCH({p},{q}) — {self.name}")
        am = arch_model(self.returns * 100, p=p, q=q,
                        mean="Constant", vol="GARCH", dist="Normal")
        res = am.fit(disp="off")
        print(res.summary())
        cond_vol = res.conditional_volatility * np.sqrt(252)
        rv = self.returns.rolling(self.window).std() * np.sqrt(252) * 100
        fig, axes = plt.subplots(2, 1, figsize=figsize, sharex=True)
        axes[0].plot(cond_vol, color="#FF5722", lw=1.2, label=f"GARCH({p},{q})")
        axes[0].plot(rv, color="#2196F3", lw=1, alpha=0.6, label=f"Rolling {self.window}gg")
        axes[0].set_title(f"Volatilità Condizionale — {self.name}"); axes[0].legend()
        axes[1].plot(res.std_resid, lw=0.8, color="#9C27B0", alpha=0.7)
        axes[1].axhline(0, color="k", lw=0.5)
        axes[1].axhline(2, color="r", lw=1, linestyle="--", alpha=0.5)
        axes[1].axhline(-2, color="r", lw=1, linestyle="--", alpha=0.5)
        axes[1].set_title("Residui Standardizzati GARCH")
        plt.tight_layout(); plt.show()
        return res, cond_vol

    def regime_detection(self, figsize=(14, 6)):
        rv = self.returns.rolling(self.window).std() * np.sqrt(252) * 100
        rv.dropna(inplace=True)
        q33, q66 = rv.quantile(0.33), rv.quantile(0.66)
        regime = pd.cut(rv, bins=[-np.inf, q33, q66, np.inf], labels=["Bassa", "Media", "Alta"])
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=figsize, sharex=True)
        for lbl, col in {"Bassa": "#4CAF50", "Media": "#FF9800", "Alta": "#F44336"}.items():
            mask = regime == lbl
            ax1.fill_between(rv.index, np.where(mask, rv, np.nan), alpha=0.7, color=col, label=lbl)
        ax1.plot(rv, lw=0.8, color="k", alpha=0.4)
        ax1.set_title(f"Regime Volatilità — {self.name}"); ax1.legend()
        rn = regime.map({"Bassa": 0, "Media": 1, "Alta": 2})
        ax2.fill_between(rn.index, rn, alpha=0.3, color="#607D8B")
        ax2.set_yticks([0, 1, 2]); ax2.set_yticklabels(["Bassa", "Media", "Alta"])
        plt.tight_layout(); plt.show()
        dist = regime.value_counts(normalize=True) * 100
        print("\n📊 Regimi:"); [print(f"   {r}: {p:.1f}%") for r, p in dist.items()]
        return regime


# ══════════════════════════════════════════════════════════════════════════════
# [H] CONFRONTO 3 PORTAFOGLI — Rolling Alpha e TEV
#     Dal video: P1=Benchmark · P2=Max-Sharpe · P3=Style-Analysis
#                Rolling window 250 giorni, TEV target 2-4%
# ══════════════════════════════════════════════════════════════════════════════

class PortfolioComparison:
    """
    Confronto 3 portafogli con:
      - Performance cumulativa e drawdown
      - Rolling Sharpe (250gg)
      - Rolling Alpha vs benchmark
      - Tracking Error Volatility con banda target 2-4%
      - Tabella riepilogativa
    """

    def __init__(self, prices, weights_p1, weights_p2, weights_p3,
                 names=("Equal-Weight (P1)", "Max-Sharpe (P2)", "Min-Var (P3)"),
                 rf=0.045):
        self.prices = prices
        self.returns = compute_returns(prices)
        self.w = [np.array(w) for w in [weights_p1, weights_p2, weights_p3]]
        self.names = names
        self.rf = rf
        self.port_rets = [
            pd.Series(self.returns[prices.columns].values @ w,
                      index=self.returns.index)
            for w in self.w]

    def _rolling_alpha(self, tgt, bm, window):
        return (tgt - bm).rolling(window).mean() * 252

    def _tev(self, tgt, bm, window):
        return (tgt - bm).rolling(window).std() * np.sqrt(252) * 100

    def plot(self, window=250, figsize=(16, 14)):
        fig = plt.figure(figsize=figsize)
        gs = gridspec.GridSpec(4, 2, figure=fig, hspace=0.55, wspace=0.3)
        cols = ["#607D8B", "#4CAF50", "#FF9800"]
        styles = ["-", "--", "-."]

        # Performance cumulativa
        ax1 = fig.add_subplot(gs[0, :])
        for ret, name, col, ls in zip(self.port_rets, self.names, cols, styles):
            ax1.plot(np.cumsum(ret), lw=2, color=col, label=name, linestyle=ls)
        ax1.set_title("Performance Cumulativa — 3 Portafogli")
        ax1.legend(fontsize=9); ax1.set_ylabel("Rendimento Log Cumulativo")

        # Drawdown
        ax2 = fig.add_subplot(gs[1, :])
        for ret, name, col in zip(self.port_rets, self.names, cols):
            cr = np.exp(np.cumsum(ret))
            dd = (cr - cr.cummax()) / cr.cummax() * 100
            ax2.plot(dd, lw=1.2, color=col, label=name)
        ax2.set_title("Drawdown dal Massimo (%)"); ax2.legend(fontsize=9); ax2.set_ylabel("Drawdown %")

        # Rolling Sharpe
        ax3 = fig.add_subplot(gs[2, 0])
        for ret, name, col in zip(self.port_rets, self.names, cols):
            rm = ret.rolling(window).mean() * 252
            rv = ret.rolling(window).std() * np.sqrt(252)
            ax3.plot((rm - self.rf) / rv, lw=1.2, color=col, label=name)
        ax3.axhline(1.0, color="k", lw=1, linestyle="--", alpha=0.4)
        ax3.axhline(0, color="r", lw=0.8, linestyle="--", alpha=0.4)
        ax3.set_title(f"Rolling Sharpe Ratio ({window}gg)"); ax3.legend(fontsize=8)

        # Rolling Alpha P2, P3 vs P1
        ax4 = fig.add_subplot(gs[2, 1])
        alpha2 = self._rolling_alpha(self.port_rets[1], self.port_rets[0], window)
        alpha3 = self._rolling_alpha(self.port_rets[2], self.port_rets[0], window)
        ax4.plot(alpha2, lw=1.2, color=cols[1], label=f"{self.names[1]} α vs P1")
        ax4.plot(alpha3, lw=1.2, color=cols[2], label=f"{self.names[2]} α vs P1")
        ax4.axhline(0, color="k", lw=0.8, linestyle="--")
        ax4.fill_between(alpha2.index, alpha2, 0, where=alpha2 > 0, alpha=0.15, color=cols[1])
        ax4.fill_between(alpha2.index, alpha2, 0, where=alpha2 < 0, alpha=0.15, color="#F44336")
        ax4.set_title(f"Rolling Alpha vs Benchmark ({window}gg)")
        ax4.set_ylabel("Alpha annualizzato"); ax4.legend(fontsize=8)

        # Tracking Error Volatility
        ax5 = fig.add_subplot(gs[3, 0])
        tev2 = self._tev(self.port_rets[1], self.port_rets[0], window)
        tev3 = self._tev(self.port_rets[2], self.port_rets[0], window)
        ax5.plot(tev2, lw=1.2, color=cols[1], label=f"TEV {self.names[1]}")
        ax5.plot(tev3, lw=1.2, color=cols[2], label=f"TEV {self.names[2]}")
        ax5.axhspan(2, 4, alpha=0.12, color="g", label="Target 2-4%")
        ax5.set_title(f"Tracking Error Volatility ({window}gg) — Target 2-4%")
        ax5.set_ylabel("TEV % ann."); ax5.legend(fontsize=8)

        # Tabella statistiche
        ax6 = fig.add_subplot(gs[3, 1])
        ax6.axis("off")
        rows_st = []
        for ret, name in zip(self.port_rets, self.names):
            ar = ret.mean() * 252 * 100
            av = ret.std() * np.sqrt(252) * 100
            sr = (ar - self.rf * 100) / av
            cr = (np.exp(np.cumsum(ret)) - 1).iloc[-1] * 100
            cum_s = pd.Series(np.exp(np.cumsum(ret)))
            mdd = ((cum_s - cum_s.cummax()) / cum_s.cummax()).min() * 100
            rows_st.append([name[:18], f"{ar:.1f}%", f"{av:.1f}%",
                            f"{sr:.2f}", f"{cr:.1f}%", f"{mdd:.1f}%"])
        tbl = ax6.table(cellText=rows_st,
                        colLabels=["Portafoglio", "Ret/Y", "Vol/Y", "SR", "Cum.", "MaxDD"],
                        loc="center", cellLoc="center")
        tbl.auto_set_font_size(False); tbl.set_fontsize(8); tbl.scale(1, 1.6)
        ax6.set_title("Riepilogo Statistiche", fontsize=10)

        plt.suptitle("Confronto 3 Portafogli — Dashboard Finale", fontsize=14, y=1.01)
        plt.tight_layout(); plt.show()


# ══════════════════════════════════════════════════════════════════════════════
# [I] STOCK SCREENER
# ══════════════════════════════════════════════════════════════════════════════

class StockScreener:
    def __init__(self, universe_tickers, start, end, rf=0.045):
        self.tickers = universe_tickers
        self.start = start; self.end = end; self.rf = rf
        self.prices = None; self.returns = None; self.scores = None

    def load_data(self):
        print(f"\n📥 Screening: {len(self.tickers)} titoli")
        self.prices = load_prices(self.tickers, self.start, self.end)
        self.returns = compute_returns(self.prices)
        return self

    def compute_factors(self):
        p = self.prices; r = self.returns; rows = []
        for ticker in self.tickers:
            if ticker not in p.columns: continue
            ps = p[ticker].dropna(); rs = r[ticker].dropna()
            if len(ps) < 252: continue
            pn = ps.iloc[-1]
            def mom(n): return (pn / ps.iloc[-n] - 1) if len(ps) >= n else np.nan
            def vol(n): return rs.iloc[-n:].std() * np.sqrt(252) if len(rs) >= n else np.nan
            ret1y = rs.iloc[-252:].mean() * 252 if len(rs) >= 252 else np.nan
            v1y = vol(252)
            sh = (ret1y - self.rf) / v1y if (v1y and v1y > 0) else np.nan
            ma200 = ps.rolling(200).mean().iloc[-1]
            dd = ((ps - ps.cummax()) / ps.cummax()).iloc[-1]
            rows.append({
                "Ticker": ticker, "Prezzo": round(pn, 2),
                "Mom_1M%": round(mom(21) * 100, 2) if not np.isnan(mom(21)) else np.nan,
                "Mom_3M%": round(mom(63) * 100, 2) if not np.isnan(mom(63)) else np.nan,
                "Mom_6M%": round(mom(126) * 100, 2) if not np.isnan(mom(126)) else np.nan,
                "Mom_12M%": round(mom(252) * 100, 2) if not np.isnan(mom(252)) else np.nan,
                "Vol_1Y%": round(v1y * 100, 2) if not np.isnan(v1y) else np.nan,
                "Sharpe_1Y": round(sh, 3) if not np.isnan(sh) else np.nan,
                "SopraMa200": 1 if pn > ma200 else 0,
                "Drawdown%": round(dd * 100, 2)})
        df = pd.DataFrame(rows)
        for col, sign in {"Mom_3M%": 1, "Mom_6M%": 1, "Mom_12M%": 1,
                          "Sharpe_1Y": 1, "Vol_1Y%": -1, "Drawdown%": -1}.items():
            if col in df.columns:
                rank = df[col].rank(pct=True, na_option="bottom")
                df[f"Score_{col}"] = rank if sign > 0 else (1 - rank)
        sc = [c for c in df.columns if c.startswith("Score_")]
        df["Score_Composito"] = (df[sc].mean(axis=1) * 100).round(1)
        df["Rank"] = df["Score_Composito"].rank(ascending=False).astype(int)
        self.scores = df.sort_values("Score_Composito", ascending=False)
        return self.scores

    def plot_screening(self, top_n=15, figsize=(16, 10)):
        if self.scores is None: self.compute_factors()
        df = self.scores.head(top_n).copy()
        fig = plt.figure(figsize=figsize)
        gs = gridspec.GridSpec(2, 2, figure=fig, hspace=0.45, wspace=0.35)
        ax1 = fig.add_subplot(gs[0, :])
        ax1.barh(df["Ticker"][::-1], df["Score_Composito"][::-1],
                 color=[COLORS[i % len(COLORS)] for i in range(len(df))][::-1])
        ax1.set_title(f"Top {top_n} — Score Composito")
        for i, (v, t) in enumerate(zip(df["Score_Composito"][::-1], df["Ticker"][::-1])):
            ax1.text(v + 0.5, i, f"{v:.1f}", va="center", fontsize=8)
        ax2 = fig.add_subplot(gs[1, 0])
        ax2.scatter(df["Mom_3M%"], df["Sharpe_1Y"],
                    c=df["Score_Composito"], cmap="RdYlGn", s=120, edgecolors="k")
        for _, row in df.iterrows():
            ax2.annotate(row["Ticker"], (row["Mom_3M%"], row["Sharpe_1Y"]),
                         textcoords="offset points", xytext=(5, 5), fontsize=7)
        ax2.axhline(0, color="k", lw=0.7, linestyle="--")
        ax2.axvline(0, color="k", lw=0.7, linestyle="--")
        ax2.set_xlabel("Momentum 3M %"); ax2.set_ylabel("Sharpe 1Y"); ax2.set_title("Mom vs Sharpe")
        ax3 = fig.add_subplot(gs[1, 1])
        ax3.scatter(df["Vol_1Y%"], df["Drawdown%"],
                    c=df["Score_Composito"], cmap="RdYlGn", s=120, edgecolors="k")
        for _, row in df.iterrows():
            ax3.annotate(row["Ticker"], (row["Vol_1Y%"], row["Drawdown%"]),
                         textcoords="offset points", xytext=(5, 5), fontsize=7)
        ax3.set_xlabel("Volatilità 1Y %"); ax3.set_ylabel("Drawdown %"); ax3.set_title("Vol vs DD")
        plt.suptitle("Stock Screening Dashboard", fontsize=14); plt.show()
        dcols = ["Rank", "Ticker", "Score_Composito", "Mom_1M%", "Mom_3M%",
                 "Sharpe_1Y", "Vol_1Y%", "Drawdown%", "SopraMa200"]
        print(self.scores[[c for c in dcols if c in self.scores.columns]
                          ].head(top_n).to_string(index=False))
        return self.scores

    def export_csv(self, path="screening_results.csv"):
        if self.scores is not None:
            self.scores.to_csv(path, index=False)
            print(f"✅ CSV: {path}")


# ══════════════════════════════════════════════════════════════════════════════
# [L] EXPORT EXCEL MULTI-FOGLIO
# ══════════════════════════════════════════════════════════════════════════════

def create_excel_dashboard(filename="Dashboard_Finanziaria.xlsx", **kw):
    wb = Workbook()
    hf = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    tfont = Font(bold=True, size=14, color="1F4E78")
    brd = Border(left=Side(style='thin'), right=Side(style='thin'),
                 top=Side(style='thin'), bottom=Side(style='thin'))

    def _write(ws, df, row0):
        for ri, row in enumerate(dataframe_to_rows(df, index=False, header=True), row0):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                if ri == row0:
                    c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center")
                c.border = brd

    def _fig2img(fig, w=700, h=300):
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
        buf.seek(0); plt.close(fig)
        img = XLImage(buf); img.width = w; img.height = h
        return img

    def _autofit(ws):
        for ci in range(1, ws.max_column + 1):
            cl = get_column_letter(ci)
            ml = max((len(str(c.value)) for c in ws[cl] if c.value), default=0)
            ws.column_dimensions[cl].width = min(ml + 2, 50)

    ef = kw.get("ef"); ms = kw.get("ms_port"); mv = kw.get("mv_port")
    replicator = kw.get("replicator"); opt_res = kw.get("opt_res")
    ols_res = kw.get("ols_res"); ridge_res = kw.get("ridge_res")
    returns_stoch = kw.get("returns_stoch"); prices_stoch = kw.get("prices_stoch")
    roll_vol = kw.get("roll_vol"); roll_sharpe = kw.get("roll_sharpe")
    cond_vol = kw.get("cond_vol")
    screening_results = kw.get("screening_results"); screener = kw.get("screener")
    diag_dummy = kw.get("diag_dummy"); diag_style = kw.get("diag_style")

    # Foglio 1: Frontiera
    ws1 = wb.active; ws1.title = "Frontiera Efficiente"
    ws1["A1"] = "FRONTIERA EFFICIENTE — Riepilogo"; ws1["A1"].font = tfont
    ws1.merge_cells("A1:E1")
    if ms and mv:
        _write(ws1, pd.DataFrame({
            "Portafoglio": ["Max Sharpe", "Min Varianza"],
            "Rendimento_%": [f"{ms['return']*100:.2f}", f"{mv['return']*100:.2f}"],
            "Volatilità_%": [f"{ms['volatility']*100:.2f}", f"{mv['volatility']*100:.2f}"],
            "Sharpe": [f"{ms['sharpe']:.3f}", f"{mv['sharpe']:.3f}"]}), 3)
    if ef and ms:
        ws1["A7"] = "PESI MAX SHARPE"; ws1["A7"].font = Font(bold=True, size=12, color="1F4E78")
        _write(ws1, pd.DataFrame({
            "Ticker": ef.tickers, "Peso_%": (ms["weights"] * 100).round(2),
            "Ret_ann_%": (ef.mu * 100).round(2),
            "Vol_ann_%": (np.sqrt(np.diag(ef.cov)) * 100).round(2)
        }).query("`Peso_%` > 0.1").sort_values("Peso_%", ascending=False), 9)
        fig, axes = plt.subplots(1, 2, figsize=(14, 5))
        sc = axes[0].scatter(ef.sim_results["Volatility"], ef.sim_results["Return"],
                             c=ef.sim_results["Sharpe"], cmap="RdYlGn", alpha=0.4, s=8)
        fc = ef.compute_frontier_curve()
        if len(fc): axes[0].plot(fc[:, 1], fc[:, 0], "b-", lw=2.5, label="Frontiera")
        axes[0].scatter(ms["volatility"], ms["return"], marker="*", s=300, c="gold",
                        edgecolors="k", label="Max Sharpe")
        axes[0].scatter(mv["volatility"], mv["return"], marker="D", s=150, c="cyan",
                        edgecolors="k", label="Min Var")
        axes[0].legend(fontsize=8); axes[0].set_title("Frontiera Efficiente")
        axes[1].pie(ms["weights"][ms["weights"] > 0.005],
                    labels=[ef.tickers[i] for i in range(ef.n) if ms["weights"][i] > 0.005],
                    autopct="%1.1f%%", startangle=140)
        axes[1].set_title("Pesi Max Sharpe")
        plt.tight_layout(); ws1.add_image(_fig2img(fig, 800, 350), "G3")

    # Foglio 2: ETF
    ws2 = wb.create_sheet("Replica ETF"); ws2["A1"] = "REPLICA ETF"; ws2["A1"].font = tfont
    ws2.merge_cells("A1:D1")
    if ols_res and ridge_res and opt_res:
        _write(ws2, pd.DataFrame({
            "Metodo": ["OLS", "Ridge", "Min-TE"],
            "TE_%": [ols_res["tracking_error"], ridge_res["tracking_error"], opt_res["tracking_error"]],
            "R²": [ols_res["r2"], ridge_res["r2"], opt_res["r2"]]}), 3)
    if replicator and opt_res:
        ws2["A8"] = "PESI MIN-TE"; ws2["A8"].font = Font(bold=True, size=12, color="1F4E78")
        _write(ws2, pd.DataFrame({
            "Strumento": replicator.universe_names,
            "Peso_%": (opt_res["weights"] * 100).round(2)
        }).sort_values("Peso_%", ascending=False), 10)
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(np.cumsum(replicator.target), label="Originale", lw=2, color="#2196F3")
        ax.plot(np.cumsum(opt_res["fitted"]), label="Replica", lw=2, color="#FF5722", linestyle="--")
        ax.set_title(f"TE:{opt_res['tracking_error']:.2f}%  R²:{opt_res['r2']:.4f}"); ax.legend()
        plt.tight_layout(); ws2.add_image(_fig2img(fig, 600, 260), "E3")

    # Foglio 3: Diagnostica
    ws3 = wb.create_sheet("Diagnostica Statistica"); ws3["A1"] = "DIAGNOSTICA STATISTICA"
    ws3["A1"].font = tfont; ws3.merge_cells("A1:G1")
    rows_d = []
    for label, diag in [("Dummy Outlier", diag_dummy), ("Style Analysis", diag_style)]:
        if diag:
            rows_d.append([label,
                           f"{diag.get('r2', 0)*100:.2f}%",
                           f"{diag.get('adj_r2', 0)*100:.2f}%",
                           f"{diag.get('durbin_watson', 0):.3f}",
                           f"{diag.get('jb_p', 0):.4f}",
                           f"{diag.get('bp_p', float('nan')):.4f}",
                           f"{diag.get('aic', 0):.1f}",
                           f"{diag.get('bic', 0):.1f}"])
    if rows_d:
        _write(ws3, pd.DataFrame(rows_d,
               columns=["Modello", "R²", "Adj R²", "DW", "JB p", "BP p", "AIC", "BIC"]), 3)

    # Foglio 4: Volatilità
    ws4 = wb.create_sheet("Volatilità"); ws4["A1"] = "ANALISI VOLATILITÀ"
    ws4["A1"].font = tfont; ws4.merge_cells("A1:D1")
    if roll_vol is not None:
        _write(ws4, pd.DataFrame({
            "Metrica": ["Media Vol (%)", "Max Vol (%)", "Min Vol (%)", "Sharpe medio"],
            "Valore": [roll_vol.mean(), roll_vol.max(), roll_vol.min(),
                       roll_sharpe.mean() if roll_sharpe is not None else "—"],
            "Data": ["—", str(roll_vol.idxmax().date()), str(roll_vol.idxmin().date()), "—"]}), 3)
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(roll_vol, color="#FF5722", lw=1.2)
        ax.fill_between(roll_vol.index, roll_vol, alpha=0.3, color="#FF5722")
        ax.set_title("Volatilità Rolling")
        plt.tight_layout(); ws4.add_image(_fig2img(fig, 600, 240), "F3")

    # Foglio 5: Screening
    ws5 = wb.create_sheet("Screening Titoli")
    ws5["A1"] = f"STOCK SCREENING — Top 15{' | '+str(len(screener.tickers))+' titoli' if screener else ''}"
    ws5["A1"].font = tfont; ws5.merge_cells("A1:L1")
    if screening_results is not None and len(screening_results) > 0:
        cols = ["Rank", "Ticker", "Score_Composito", "Prezzo",
                "Mom_1M%", "Mom_3M%", "Mom_6M%", "Mom_12M%",
                "Sharpe_1Y", "Vol_1Y%", "Drawdown%", "SopraMa200"]
        _write(ws5, screening_results[[c for c in cols if c in screening_results.columns]].head(15), 3)
        t10 = screening_results.head(10)
        fig, ax = plt.subplots(figsize=(8, 6))
        sc3 = ax.scatter(t10["Mom_3M%"], t10["Sharpe_1Y"],
                         c=t10["Score_Composito"], cmap="RdYlGn", s=150, edgecolors="k")
        for _, row in t10.iterrows():
            ax.annotate(row["Ticker"], (row["Mom_3M%"], row["Sharpe_1Y"]),
                        textcoords="offset points", xytext=(5, 5), fontsize=8)
        ax.axhline(0, color="k", lw=0.7, linestyle="--")
        ax.axvline(0, color="k", lw=0.7, linestyle="--")
        ax.set_title("Top 10: Mom 3M vs Sharpe"); plt.colorbar(sc3, label="Score")
        plt.tight_layout(); ws5.add_image(_fig2img(fig, 500, 380), "N3")

    # ── Foglio 6: Confronto 3 Portafogli ─────────────────────────────────────
    pc_obj     = kw.get("pc_obj")       # PortfolioComparison instance
    roll_window = kw.get("roll_window", 250)

    ws6 = wb.create_sheet("Confronto Portafogli")
    ws6["A1"] = "CONFRONTO 3 PORTAFOGLI — Performance · Rolling Alpha · TEV"
    ws6["A1"].font = tfont; ws6.merge_cells("A1:G1")

    if pc_obj is not None:
        # Tabella statistiche aggregate
        rows_pc = []
        for ret, name in zip(pc_obj.port_rets, pc_obj.names):
            ar  = ret.mean() * 252 * 100
            av  = ret.std()  * np.sqrt(252) * 100
            sr  = (ar - pc_obj.rf * 100) / av
            cr  = (np.exp(np.cumsum(ret)) - 1).iloc[-1] * 100
            cum_s   = pd.Series(np.exp(np.cumsum(ret)))
            mdd = ((cum_s - cum_s.cummax()) / cum_s.cummax()).min() * 100
            cal = ar / abs(mdd) if mdd != 0 else np.nan  # Calmar ratio
            rows_pc.append([name, f"{ar:.2f}%", f"{av:.2f}%",
                            f"{sr:.3f}", f"{cr:.1f}%",
                            f"{mdd:.1f}%", f"{cal:.3f}"])
        df_pc = pd.DataFrame(rows_pc,
                             columns=["Portafoglio","Ret/Y","Vol/Y",
                                      "Sharpe","Cumulativo","MaxDD","Calmar"])
        _write(ws6, df_pc, 3)

        ws6["A9"]  = "ROLLING ALPHA vs P1 (Equal-Weight)"; ws6["A9"].font  = Font(bold=True, size=12, color="1F4E78")
        ws6["A10"] = f"Finestra rolling: {roll_window} giorni"

        p_colors = ["#607D8B", "#4CAF50", "#FF9800"]
        styles   = ["-", "--", "-."]

        # ── Fig A: performance cumulativa ───────────────────────────────────
        fig_a, ax_a = plt.subplots(figsize=(12, 4))
        for ret, name, col, ls in zip(pc_obj.port_rets, pc_obj.names, p_colors, styles):
            ax_a.plot(np.cumsum(ret), lw=2, color=col, label=name, linestyle=ls)
        ax_a.set_title("Performance Cumulativa (Log-Return)")
        ax_a.legend(fontsize=9); ax_a.set_ylabel("Rend. Log Cumulativo")
        plt.tight_layout()
        ws6.add_image(_fig2img(fig_a, 680, 260), "A12")

        # ── Fig B: rolling alpha P2, P3 vs P1 ──────────────────────────────
        alpha2 = (pc_obj.port_rets[1] - pc_obj.port_rets[0]).rolling(roll_window).mean() * 252
        alpha3 = (pc_obj.port_rets[2] - pc_obj.port_rets[0]).rolling(roll_window).mean() * 252

        fig_b, ax_b = plt.subplots(figsize=(12, 4))
        ax_b.plot(alpha2, lw=1.5, color=p_colors[1], label=f"{pc_obj.names[1]} α vs P1")
        ax_b.plot(alpha3, lw=1.5, color=p_colors[2], label=f"{pc_obj.names[2]} α vs P1")
        ax_b.axhline(0, color="k", lw=0.8, linestyle="--")
        ax_b.fill_between(alpha2.index, alpha2, 0, where=alpha2 > 0, alpha=0.15, color=p_colors[1])
        ax_b.fill_between(alpha2.index, alpha2, 0, where=alpha2 < 0, alpha=0.15, color="#F44336")
        ax_b.set_title(f"Rolling Alpha vs Benchmark ({roll_window}gg)")
        ax_b.set_ylabel("Alpha annualizzato"); ax_b.legend(fontsize=9)
        plt.tight_layout()
        ws6.add_image(_fig2img(fig_b, 680, 260), "A33")

        # ── Fig C: Tracking Error Volatility ────────────────────────────────
        tev2 = (pc_obj.port_rets[1] - pc_obj.port_rets[0]).rolling(roll_window).std() * np.sqrt(252) * 100
        tev3 = (pc_obj.port_rets[2] - pc_obj.port_rets[0]).rolling(roll_window).std() * np.sqrt(252) * 100

        fig_c, ax_c = plt.subplots(figsize=(12, 4))
        ax_c.plot(tev2, lw=1.5, color=p_colors[1], label=f"TEV {pc_obj.names[1]}")
        ax_c.plot(tev3, lw=1.5, color=p_colors[2], label=f"TEV {pc_obj.names[2]}")
        ax_c.axhspan(2, 4, alpha=0.12, color="g", label="Target 2-4%")
        ax_c.set_title(f"Tracking Error Volatility — Banda target 2-4% ({roll_window}gg)")
        ax_c.set_ylabel("TEV % ann."); ax_c.legend(fontsize=9)
        plt.tight_layout()
        ws6.add_image(_fig2img(fig_c, 680, 260), "A53")

        # ── Fig D: Drawdown ─────────────────────────────────────────────────
        fig_d, ax_d = plt.subplots(figsize=(12, 4))
        for ret, name, col, ls in zip(pc_obj.port_rets, pc_obj.names, p_colors, styles):
            cr = np.exp(np.cumsum(ret))
            dd = (cr - cr.cummax()) / cr.cummax() * 100
            ax_d.plot(dd, lw=1.5, color=col, label=name, linestyle=ls)
        ax_d.set_title("Drawdown dal Massimo (%)"); ax_d.set_ylabel("Drawdown %")
        ax_d.legend(fontsize=9)
        plt.tight_layout()
        ws6.add_image(_fig2img(fig_d, 680, 260), "A73")

        # Alpha summary stats
        ws6["A93"] = "RIEPILOGO ALPHA (media periodo)"; ws6["A93"].font = Font(bold=True, size=11, color="1F4E78")
        alpha_rows = []
        for name, alpha in [(pc_obj.names[1], alpha2), (pc_obj.names[2], alpha3)]:
            alpha_rows.append([
                name,
                f"{alpha.mean()*100:.2f}%",
                f"{(alpha > 0).mean()*100:.1f}%",
                f"{alpha.max()*100:.2f}%",
                f"{alpha.min()*100:.2f}%",
            ])
        _write(ws6, pd.DataFrame(alpha_rows,
                                  columns=["Portafoglio", "Alpha medio",
                                           "% gg positivi", "Alpha max", "Alpha min"]), 94)

    # ── Foglio 7: ARIMA Forecast ──────────────────────────────────────────────
    mu_arima_kw = kw.get("mu_arima")   # pd.Series ticker→annualised forecast
    ef_obj      = kw.get("ef")

    ws7 = wb.create_sheet("ARIMA Forecast")
    ws7["A1"] = "ARIMA FORECAST — Rendimenti Attesi Storici vs Previsti"
    ws7["A1"].font = tfont; ws7.merge_cells("A1:F1")

    if mu_arima_kw is not None and ef_obj is not None:
        mu_hist = ef_obj.mu  # annualised historical μ
        tickers_arima = [t for t in mu_arima_kw.index if t in mu_hist.index]

        rows_arima = []
        for t in tickers_arima:
            mh = float(mu_hist.loc[t]) * 100
            mf = float(mu_arima_kw.loc[t]) * 100
            diff = mf - mh
            sig  = "📈 RIALZO" if diff > 1 else ("📉 RIBASSO" if diff < -1 else "➡️  STABILE")
            rows_arima.append([t, f"{mh:.2f}%", f"{mf:.2f}%",
                               f"{diff:+.2f}%", sig])

        df_arima = pd.DataFrame(rows_arima,
                                columns=["Ticker", "μ Storico %",
                                         "μ ARIMA %", "Δ %", "Segnale"])
        _write(ws7, df_arima, 3)

        # ── Bar chart confronto ──────────────────────────────────────────────
        mu_h_vals = [float(mu_hist.loc[t]) * 100 for t in tickers_arima]
        mu_f_vals = [float(mu_arima_kw.loc[t]) * 100 for t in tickers_arima]
        x = np.arange(len(tickers_arima)); w = 0.35

        fig_e, ax_e = plt.subplots(figsize=(12, 5))
        ax_e.bar(x - w/2, mu_h_vals, w, label="μ Storico %",       color="#2196F3", alpha=0.85)
        ax_e.bar(x + w/2, mu_f_vals, w, label="μ ARIMA Forecast %", color="#FF5722", alpha=0.85)
        ax_e.set_xticks(x); ax_e.set_xticklabels(tickers_arima, rotation=45, ha="right")
        ax_e.axhline(0, color="k", lw=0.8, linestyle="--")
        ax_e.set_ylabel("Rendimento % annualizzato")
        ax_e.set_title("Confronto μ Storico vs ARIMA Forecast per Asset")
        ax_e.legend(); plt.tight_layout()
        ws7.add_image(_fig2img(fig_e, 680, 300), "A20")

        # Nota metodologica
        ws7["A40"] = "NOTE METODOLOGICHE"
        ws7["A40"].font = Font(bold=True, size=11, color="1F4E78")
        notes = [
            ["ARIMA(p,d,q)", "Autoregressive Integrated Moving Average — modella trend, stagionalità ed errore"],
            ["μ ARIMA",      "Rendimento medio annualizzato della previsione ARIMA sul periodo horizon"],
            ["μ Storico",    "Rendimento medio annualizzato calcolato sull'intero campione in-sample"],
            ["Segnale",      "📈 RIALZO se μ_ARIMA > μ_Storico + 1pp  |  📉 RIBASSO se < -1pp"],
            ["Uso",          "μ ARIMA usato come override nel ricalcolo della Frontiera Efficiente"],
        ]
        for i, (k, v) in enumerate(notes):
            ws7.cell(row=41+i, column=1, value=k).font = Font(bold=True)
            ws7.cell(row=41+i, column=2, value=v)

    # ── Foglio 0: Allocazione Operativa (tutti e 3 i portafogli) ─────────────
    # Creato per ultimo e poi spostato in prima posizione
    ws0 = wb.create_sheet("📋 Allocazione Operativa")

    # Stili dedicati
    green_fill  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    blue_fill   = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    orange_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    bold14      = Font(bold=True, size=14, color="1F4E78")
    bold11      = Font(bold=True, size=11)

    ws0["A1"] = "📋 GUIDA OPERATIVA — Cosa comprare per ciascun portafoglio"
    ws0["A1"].font = tfont; ws0.merge_cells("A1:L1")

    # ── Sottotitoli dei 3 portafogli ─────────────────────────────────────────
    ws0["A3"] = "P1 — Equal-Weight"
    ws0["A3"].font = Font(bold=True, size=13, color="607D8B")
    ws0["A3"].fill = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
    ws0.merge_cells("A3:C3")
    ws0["A4"] = "Strategia: stessa cifra su tutto. Zero ottimizzazione."
    ws0["A4"].font = Font(italic=True, size=10, color="607D8B")
    ws0.merge_cells("A4:C4")

    ws0["E3"] = "P2 — Max-Sharpe  ⭐ CONSIGLIATO"
    ws0["E3"].font = Font(bold=True, size=13, color="2E7D32")
    ws0["E3"].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    ws0.merge_cells("E3:G3")
    ws0["E4"] = "Strategia: massimo guadagno per ogni euro di rischio (Sharpe ottimizzato)."
    ws0["E4"].font = Font(italic=True, size=10, color="2E7D32")
    ws0.merge_cells("E4:G4")

    ws0["I3"] = "P3 — Min-Varianza"
    ws0["I3"].font = Font(bold=True, size=13, color="E65100")
    ws0["I3"].fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    ws0.merge_cells("I3:K3")
    ws0["I4"] = "Strategia: oscillazioni minime. Guadagna meno ma cade meno."
    ws0["I4"].font = Font(italic=True, size=10, color="E65100")
    ws0.merge_cells("I4:K4")

    if ef and ms and mv and pc_obj is not None:
        tickers     = ef.tickers
        n_assets    = len(tickers)
        w_p1        = pc_obj.w[0]   # equal-weight
        w_p2        = pc_obj.w[1]   # max-sharpe
        w_p3        = pc_obj.w[2]   # min-var

        # ── Intestazioni colonne tabella confronto ────────────────────────────
        headers = [
            ("A", "Ticker"),
            ("B", "P1 Peso %"),   ("C", "P1 su €100k"),
            ("E", "P2 Peso %"),   ("F", "P2 su €100k"),   ("G", "P2 Differenza vs P1"),
            ("I", "P3 Peso %"),   ("J", "P3 su €100k"),   ("K", "P3 Differenza vs P1"),
        ]
        for col, label in headers:
            c = ws0[f"{col}6"]
            c.value = label
            c.font  = hfont
            c.fill  = hf
            c.alignment = Alignment(horizontal="center")
            c.border = brd

        # ── Righe per ogni titolo ─────────────────────────────────────────────
        for i, ticker in enumerate(tickers):
            row = 7 + i
            wp1 = float(w_p1[i]) * 100
            wp2 = float(w_p2[i]) * 100
            wp3 = float(w_p3[i]) * 100
            diff_p2 = wp2 - wp1
            diff_p3 = wp3 - wp1

            def _cell(col, val, fill=None, num_fmt=None):
                c = ws0.cell(row=row, column=col, value=val)
                c.border = brd
                c.alignment = Alignment(horizontal="right")
                if fill: c.fill = fill
                if num_fmt: c.number_format = num_fmt
                return c

            _cell(1, ticker).alignment = Alignment(horizontal="left")
            ws0.cell(row=row, column=1).font = Font(bold=True)

            # P1
            _cell(2, round(wp1, 2), blue_fill)
            _cell(3, round(wp1 * 1000, 0), blue_fill, '"€"#,##0')

            # P2
            _cell(5, round(wp2, 2), green_fill)
            _cell(6, round(wp2 * 1000, 0), green_fill, '"€"#,##0')
            c_d2 = _cell(7, round(diff_p2, 2), green_fill)
            c_d2.font = Font(color="2E7D32" if diff_p2 >= 0 else "C62828", bold=True)
            c_d2.value = f"{diff_p2:+.2f}%"

            # P3
            _cell(9,  round(wp3, 2), orange_fill)
            _cell(10, round(wp3 * 1000, 0), orange_fill, '"€"#,##0')
            c_d3 = _cell(11, round(diff_p3, 2), orange_fill)
            c_d3.font = Font(color="2E7D32" if diff_p3 >= 0 else "C62828", bold=True)
            c_d3.value = f"{diff_p3:+.2f}%"

        # ── Riga TOTALE ───────────────────────────────────────────────────────
        tot_row = 7 + n_assets
        for col, val, fill in [
            (2,  100.0, blue_fill),   (3,  100000, blue_fill),
            (5,  100.0, green_fill),  (6,  100000, green_fill),  (7, "—", green_fill),
            (9,  100.0, orange_fill), (10, 100000, orange_fill), (11, "—", orange_fill),
        ]:
            c = ws0.cell(row=tot_row, column=col, value=val)
            c.font   = Font(bold=True)
            c.fill   = fill
            c.border = brd
            c.alignment = Alignment(horizontal="right")
            if isinstance(val, (int, float)) and val == 100000:
                c.number_format = '"€"#,##0'
        ws0.cell(row=tot_row, column=1, value="TOTALE").font = Font(bold=True)

        # ── Grafici a torta 3 portafogli ──────────────────────────────────────
        fig_pie, axes_pie = plt.subplots(1, 3, figsize=(18, 6))
        pie_data = [
            (w_p1, "P1 — Equal-Weight",   "#607D8B"),
            (w_p2, "P2 — Max-Sharpe ⭐",  "#4CAF50"),
            (w_p3, "P3 — Min-Varianza",   "#FF9800"),
        ]
        for ax, (weights, title, accent) in zip(axes_pie, pie_data):
            mask = np.array(weights) > 0.005
            ax.pie(
                np.array(weights)[mask],
                labels=[tickers[j] for j in range(n_assets) if mask[j]],
                autopct="%1.1f%%", startangle=140,
                colors=COLORS[:mask.sum()],
                wedgeprops={"edgecolor": "white", "linewidth": 2})
            ax.set_title(title, fontsize=13, fontweight="bold", color=accent)
        plt.suptitle("Confronto Allocazione — P1 vs P2 vs P3", fontsize=14, y=1.02)
        plt.tight_layout()
        ws0.add_image(_fig2img(fig_pie, 900, 340), f"A{tot_row + 3}")

        # ── Legenda plain Italian ──────────────────────────────────────────────
        leg_row = tot_row + 25
        ws0.cell(row=leg_row, column=1,
                 value="📖 COME LEGGERE QUESTA TABELLA").font = bold14
        ws0.merge_cells(f"A{leg_row}:L{leg_row}")

        legend_rows = [
            ("Peso %",           "La percentuale del tuo capitale da investire in quel titolo."),
            ("su €100k",         "Se hai €100.000, quanti euro metti su quel titolo."),
            ("Differenza vs P1", "Quanto P2 o P3 si discostano dalla strategia 'metto tutto uguale'."),
            ("P1 Equal-Weight",  "Strategia passiva. Nessun calcolo. Stessa % su ogni titolo."),
            ("P2 Max-Sharpe ⭐", "La matematica ottimizza il rapporto guadagno/rischio. Risultato storico: +23.3%/anno, +406% in 7 anni."),
            ("P3 Min-Varianza",  "Il portafoglio che oscilla meno. Buono se hai paura dei crolli ma rende solo +8.9%/anno."),
            ("Calmar Ratio",     "Guadagno annuo diviso il peggior crollo. Più è alto, meglio è. P2 = 0.691 (il migliore)."),
        ]
        for j, (term, desc) in enumerate(legend_rows):
            r = leg_row + 2 + j
            c_t = ws0.cell(row=r, column=1, value=term)
            c_t.font = Font(bold=True, size=11)
            c_d = ws0.cell(row=r, column=2, value=desc)
            c_d.font = Font(size=11)
            ws0.merge_cells(f"B{r}:L{r}")

    # ── Foglio 8: Rischio Cambio EUR/USD ─────────────────────────────────────
    fx_analyzer = kw.get("fx_analyzer")
    ms_weights_fx = kw.get("ms_weights_fx")

    ws8 = wb.create_sheet("Rischio Cambio EUR")
    ws8["A1"] = "🇮🇹 RISCHIO CAMBIO EUR/USD — Analisi Investitore Italiano"
    ws8["A1"].font = tfont; ws8.merge_cells("A1:H1")

    if fx_analyzer is not None:
        df_fx = fx_analyzer.impact_table()
        ann_benefit = (fx_analyzer.rate_usd - fx_analyzer.rate_eur) * 100

        # Tabella impatto
        ws8["A3"] = "IMPATTO FX PER ASSET"; ws8["A3"].font = Font(bold=True, size=12, color="1F4E78")
        _write(ws8, df_fx, 5)

        # Nota interpretativa
        note_row = 5 + len(df_fx) + 2
        notes_fx = [
            ["Rend. USD %",        "Rendimento implicito dell'asset in USD (componente pura)"],
            ["Rend. EUR %",        "Rendimento che hai ricevuto in EUR su Xetra (USD + FX)"],
            ["Rend. EUR-Hed %",    "Rendimento EUR se avessi coperto il cambio con forward"],
            ["Impatto FX %",       "Quanto EUR/USD ha aggiunto o tolto al tuo rendimento"],
            ["Beneficio Hedge %",  f"Guadagno extra dall'hedging: differenziale Fed-BCE = +{ann_benefit:.2f}%/anno"],
            ["Regola pratica",     "EUR forte → impatto FX negativo. EUR debole → impatto FX positivo."],
            ["Hedging 2024-25",    f"FAVOREVOLE: Fed({fx_analyzer.rate_usd*100:.2f}%) > BCE({fx_analyzer.rate_eur*100:.2f}%) → guadagni {ann_benefit:.2f}%/anno coprendo il cambio"],
        ]
        ws8.cell(row=note_row, column=1, value="GUIDA ALLA LETTURA").font = Font(bold=True, size=12, color="1F4E78")
        for j, (term, desc) in enumerate(notes_fx):
            r = note_row + 2 + j
            ws8.cell(row=r, column=1, value=term).font = Font(bold=True)
            c = ws8.cell(row=r, column=2, value=desc)
            ws8.merge_cells(f"B{r}:H{r}")

        # Grafico EUR/USD
        if fx_analyzer.eurusd is not None:
            fig_fx, ax_fx = plt.subplots(figsize=(12, 4))
            mean_fx = fx_analyzer.eurusd.mean()
            ax_fx.plot(fx_analyzer.eurusd, color="#2196F3", lw=1.5, label="EUR/USD")
            ax_fx.fill_between(fx_analyzer.eurusd.index,
                               fx_analyzer.eurusd, mean_fx,
                               where=fx_analyzer.eurusd > mean_fx,
                               alpha=0.25, color="#F44336", label="EUR forte → penalità")
            ax_fx.fill_between(fx_analyzer.eurusd.index,
                               fx_analyzer.eurusd, mean_fx,
                               where=fx_analyzer.eurusd < mean_fx,
                               alpha=0.25, color="#4CAF50", label="EUR debole → bonus")
            ax_fx.axhline(mean_fx, color="k", lw=0.8, linestyle="--",
                          label=f"Media: {mean_fx:.4f}")
            ax_fx.set_title("EUR/USD nel periodo di analisi")
            ax_fx.legend(fontsize=8); ax_fx.set_ylabel("EUR/USD")
            plt.tight_layout()
            ws8.add_image(_fig2img(fig_fx, 700, 260), f"A{note_row + len(notes_fx) + 4}")

    # Sposta ws0 in prima posizione
    wb.move_sheet(ws0, offset=-(len(wb.sheetnames) - 1))

    for ws in [ws0, ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8]: _autofit(ws)
    wb.save(filename)
    print(f"\n✅ Excel salvato: {filename}")
    print("   Fogli: 📋 Allocazione · Frontiera · ETF · Diagnostica · Volatilità · Screening · Confronto · ARIMA · 🇮🇹 Rischio Cambio")
    return filename


# ══════════════════════════════════════════════════════════════════════════════
# [M] ANALISI RISCHIO CAMBIO EUR/USD — Investitore Italiano 🇮🇹
#
# Metodologia:
#   I titoli su Xetra (APC.DE, MSF.DE…) sono già prezzati in EUR.
#   Il rendimento EUR = rendimento USD + variazione USD/EUR
#                     = rendimento USD - variazione EUR/USD
#
#   Decomposizione:
#     r_EUR (Xetra) ≈ r_USD_asset - r_EURUSD
#     Componente asset = r_USD_implied = r_EUR + r_EURUSD
#     Componente FX    = -r_EURUSD
#     EUR forte → r_EURUSD > 0 → penalità per investitore EUR in asset USD
#     EUR debole → r_EURUSD < 0 → bonus per investitore EUR in asset USD
#
#   Hedging (contratto forward USD/EUR):
#     Costo hedge = (tasso EUR - tasso USD) / 252 per giorno
#     In 2024-25: BCE ~3.0%, Fed ~5.25% → differenziale ~-2.25%/anno
#     → BENEFICIO NETTO per l'investitore europeo che copre il cambio
#       (vende USD forward ad un tasso più alto)
#     r_EUR_hedged ≈ r_USD_implied + (rate_USD - rate_EUR)/252
# ══════════════════════════════════════════════════════════════════════════════

class FXRiskAnalyzer:
    """
    Analisi completa del rischio cambio EUR/USD per investitore italiano
    che detiene asset USA acquistati su Xetra (prezzi già in EUR).
    """

    def __init__(self, returns_eur_xetra, tickers, start, end,
                 rate_usd=0.0525, rate_eur=0.030):
        """
        returns_eur_xetra : DataFrame rendimenti log da Xetra (già in EUR)
        rate_usd          : tasso Fed annuo (default 5.25%)
        rate_eur          : tasso BCE annuo (default 3.00%)
        """
        self.ret_eur    = returns_eur_xetra
        self.tickers    = [t for t in tickers if t in returns_eur_xetra.columns]
        self.start      = start
        self.end        = end
        self.rate_usd   = rate_usd
        self.rate_eur   = rate_eur
        # Costo hedge giornaliero (negativo = beneficio per EUR investor)
        self.daily_hedge = (rate_eur - rate_usd) / 252
        self.eurusd     = None
        self.fx_ret     = None
        self.ret_usd_implied  = None
        self.ret_eur_hedged   = None

    # ── Step 1: scarica EUR/USD ───────────────────────────────────────────────
    def download_fx(self):
        print("  📥 EUR/USD (EURUSD=X)...")
        fx = yf.download("EURUSD=X", start=self.start, end=self.end,
                         auto_adjust=True, progress=False)["Close"].squeeze()
        fx.ffill(inplace=True)
        self.eurusd  = fx
        self.fx_ret  = np.log(fx / fx.shift(1)).dropna()
        print(f"  ✅ {len(self.eurusd)} giorni EUR/USD scaricati.")
        return self.eurusd

    # ── Step 2: decomponi rendimenti ─────────────────────────────────────────
    def decompose(self):
        """
        r_EUR  = r_USD - r_EURUSD  →  r_USD = r_EUR + r_EURUSD
        r_EUR_hedged ≈ r_USD + daily_hedge
                     = r_EUR + r_EURUSD + (rate_EUR - rate_USD)/252
        """
        if self.fx_ret is None:
            self.download_fx()

        idx = self.ret_eur.index.intersection(self.fx_ret.index)
        re  = self.ret_eur.loc[idx, self.tickers]
        fx  = self.fx_ret.reindex(idx).fillna(0)

        # USD implied return (ricostruito)
        self.ret_usd_implied = re.add(fx, axis=0)
        # EUR hedged: rimuove FX, aggiunge differenziale tassi
        self.ret_eur_hedged  = self.ret_usd_implied + self.daily_hedge

        return self.ret_usd_implied, self.ret_eur_hedged

    # ── Step 3: tabella impatto ───────────────────────────────────────────────
    def impact_table(self):
        if self.ret_usd_implied is None:
            self.decompose()

        rows = []
        for t in self.tickers:
            r_eur  = self.ret_eur[t].dropna()
            r_usd  = self.ret_usd_implied[t].dropna()
            r_hed  = self.ret_eur_hedged[t].dropna()

            ann_eur  = r_eur.mean()  * 252 * 100
            ann_usd  = r_usd.mean()  * 252 * 100
            ann_hed  = r_hed.mean()  * 252 * 100
            fx_drag  = ann_eur - ann_usd        # negativo se EUR forte
            hed_ben  = ann_hed - ann_eur        # beneficio hedge

            vol_eur  = r_eur.std()  * np.sqrt(252) * 100
            vol_hed  = r_hed.std()  * np.sqrt(252) * 100

            rows.append({
                "Ticker":           t,
                "Rend. USD %":      round(ann_usd,  2),
                "Rend. EUR %":      round(ann_eur,  2),
                "Rend. EUR-Hed %":  round(ann_hed,  2),
                "Impatto FX %":     round(fx_drag,  2),
                "Beneficio Hedge %":round(hed_ben,  2),
                "Vol EUR %":        round(vol_eur,  2),
                "Vol Hedged %":     round(vol_hed,  2),
            })
        return pd.DataFrame(rows)

    # ── Step 4: grafici ───────────────────────────────────────────────────────
    def plot(self, ms_weights=None, figsize=(15, 14)):
        if self.ret_usd_implied is None:
            self.decompose()

        df_imp = self.impact_table()
        fig    = plt.figure(figsize=figsize)
        gs     = gridspec.GridSpec(4, 2, figure=fig, hspace=0.55, wspace=0.35)

        # ── 1. EUR/USD nel tempo ──────────────────────────────────────────────
        ax1 = fig.add_subplot(gs[0, :])
        ax1.plot(self.eurusd, color="#2196F3", lw=1.5, label="EUR/USD")
        mean_fx = self.eurusd.mean()
        ax1.fill_between(self.eurusd.index, self.eurusd, mean_fx,
                         where=self.eurusd > mean_fx,
                         alpha=0.25, color="#F44336",
                         label="EUR forte → penalità su asset USD")
        ax1.fill_between(self.eurusd.index, self.eurusd, mean_fx,
                         where=self.eurusd < mean_fx,
                         alpha=0.25, color="#4CAF50",
                         label="EUR debole → bonus su asset USD")
        ax1.axhline(mean_fx, color="k", lw=0.8, linestyle="--",
                    label=f"Media periodo: {mean_fx:.4f}")
        ax1.set_title("EUR/USD — Quando EUR si apprezza, i tuoi asset USA valgono meno in €")
        ax1.legend(fontsize=8); ax1.set_ylabel("EUR/USD")

        # ── 2. Barre: USD vs EUR vs Hedged per asset ──────────────────────────
        ax2 = fig.add_subplot(gs[1, :])
        x = np.arange(len(df_imp)); w = 0.25
        ax2.bar(x - w,   df_imp["Rend. USD %"],     w,
                label="In USD (base)", color="#607D8B", alpha=0.85)
        ax2.bar(x,       df_imp["Rend. EUR %"],      w,
                label="In EUR (no-hedge)", color="#FF5722", alpha=0.85)
        ax2.bar(x + w,   df_imp["Rend. EUR-Hed %"],  w,
                label=f"In EUR (hedged, diff. tassi {(self.rate_usd-self.rate_eur)*100:.2f}%/y)",
                color="#4CAF50", alpha=0.85)
        ax2.set_xticks(x)
        ax2.set_xticklabels(df_imp["Ticker"], rotation=45, ha="right")
        ax2.axhline(0, color="k", lw=0.7, linestyle="--")
        ax2.set_ylabel("Rendimento % annuo")
        ax2.set_title("Impatto FX su ogni Asset — USD vs EUR no-hedge vs EUR hedged")
        ax2.legend(fontsize=8)

        # ── 3. Barre: impatto FX isolato ──────────────────────────────────────
        ax3 = fig.add_subplot(gs[2, 0])
        colors_fx = ["#4CAF50" if v >= 0 else "#F44336"
                     for v in df_imp["Impatto FX %"]]
        ax3.bar(df_imp["Ticker"], df_imp["Impatto FX %"],
                color=colors_fx, edgecolor="white")
        ax3.axhline(0, color="k", lw=0.8, linestyle="--")
        ax3.set_title("Impatto FX per asset\n(verde=EUR debole ti ha aiutato)")
        ax3.set_ylabel("pp/anno")
        ax3.tick_params(axis="x", rotation=45)

        # ── 4. Portafoglio aggregato hedged vs no-hedge ───────────────────────
        ax4 = fig.add_subplot(gs[2, 1])
        if ms_weights is not None and len(ms_weights) >= len(self.tickers):
            idx  = self.ret_eur.index.intersection(self.fx_ret.index)
            cols = self.tickers
            w_a  = np.array(ms_weights[:len(cols)])
            w_a  = w_a / w_a.sum()

            rp_eur = self.ret_eur.loc[idx, cols].values @ w_a
            rp_usd = self.ret_usd_implied.loc[idx, cols].values @ w_a
            rp_hed = self.ret_eur_hedged.loc[idx, cols].values @ w_a

            ax4.plot(np.cumsum(rp_usd), label="USD",         color="#607D8B", lw=2)
            ax4.plot(np.cumsum(rp_eur), label="EUR no-hedge",color="#FF5722", lw=2, ls="--")
            ax4.plot(np.cumsum(rp_hed), label="EUR hedged",  color="#4CAF50", lw=2, ls="-.")
            ax4.set_title("Portafoglio Max-Sharpe\nUSD vs EUR vs EUR-Hedged")
            ax4.legend(fontsize=8); ax4.set_ylabel("Log-Return Cumulativo")
        else:
            ax4.text(0.5, 0.5, "Esegui Frontiera\nper vedere portafoglio",
                     ha="center", va="center", fontsize=12,
                     transform=ax4.transAxes, color="#9E9E9E")
            ax4.set_title("Portafoglio aggregato")

        # ── 5. Volatilità EUR vs Hedged ───────────────────────────────────────
        ax5 = fig.add_subplot(gs[3, :])
        x2 = np.arange(len(df_imp)); w2 = 0.35
        ax5.bar(x2 - w2/2, df_imp["Vol EUR %"],    w2,
                label="Vol EUR no-hedge", color="#FF5722", alpha=0.8)
        ax5.bar(x2 + w2/2, df_imp["Vol Hedged %"], w2,
                label="Vol EUR hedged",   color="#4CAF50", alpha=0.8)
        ax5.set_xticks(x2)
        ax5.set_xticklabels(df_imp["Ticker"], rotation=45, ha="right")
        ax5.set_ylabel("Volatilità % annua")
        ax5.set_title("Volatilità — L'hedging riduce la volatilità eliminando il rumore FX")
        ax5.legend(fontsize=8)

        ann_hedge_benefit = (self.rate_usd - self.rate_eur) * 100
        plt.suptitle(
            f"Rischio Cambio EUR/USD — Investitore Italiano 🇮🇹\n"
            f"Differenziale tassi Fed-BCE: +{ann_hedge_benefit:.2f}%/anno "
            f"(beneficio per chi copre il cambio)",
            fontsize=13, y=1.01)
        plt.tight_layout()
        plt.show()

        print("\n📊 Tabella Impatto FX:")
        print(df_imp.to_string(index=False))
        avg_drag = df_imp["Impatto FX %"].mean()
        print(f"\n   Impatto FX medio portafoglio: {avg_drag:+.2f}%/anno")
        print(f"   Beneficio hedge medio:         {df_imp['Beneficio Hedge %'].mean():+.2f}%/anno")
        print(f"   → Differenziale Fed-BCE ({self.rate_usd*100:.2f}% - {self.rate_eur*100:.2f}%) = "
              f"+{ann_hedge_benefit:.2f}%/anno di GUADAGNO netto dalla copertura cambio")

        return df_imp



# ── Dash App ───────────────────────────────────────────────────────────────

# ── Default tickers IS20 ────────────────────────────────────────────────────
DEFAULT_TICKERS   = "NVDA,AAPL,MSFT,AMZN,GOOGL,AVGO,META,TSLA,JPM,LLY,WMT,COST,V,NFLX,MA,ABBV"
DEFAULT_BENCHMARK = "CSSPX.MI"
DEFAULT_START     = "2016-01-01"
DEFAULT_END       = "2026-04-17"
IS20_W = dict(zip(
    ["NVDA","AAPL","MSFT","AMZN","GOOGL","AVGO","META","TSLA",
     "JPM","LLY","WMT","COST","V","NFLX","MA","ABBV"],
    [0.165,0.133,0.107,0.084,0.067,0.065,0.051,0.037,
     0.029,0.025,0.021,0.017,0.018,0.016,0.014,0.013]))

COLORS_P = ["#FF6B35","#4CAF50","#2196F3","#9C27B0","#FF9800",
            "#00BCD4","#F44336","#8BC34A","#E91E63","#607D8B"]

# ── Colori tematici per eventi storici ─────────────────────────────────────
EVENTS_META = {
    "Brexit":               {"start":"2016-06-23","end":"2016-10-31","color":"rgba(255,193,7,0.15)"},
    "Guerra USA-Cina":      {"start":"2018-07-06","end":"2019-12-31","color":"rgba(255,87,34,0.15)"},
    "COVID Crash":          {"start":"2020-02-01","end":"2020-06-30","color":"rgba(244,67,54,0.20)"},
    "Ripresa Post-COVID":   {"start":"2020-07-01","end":"2021-06-30","color":"rgba(33,150,243,0.12)"},
    "Inflazione USA":       {"start":"2021-10-01","end":"2022-12-31","color":"rgba(156,39,176,0.15)"},
    "Guerra Russia-Ucraina":{"start":"2022-02-24","end":"2022-12-31","color":"rgba(255,87,34,0.18)"},
    "SVB Crisis":           {"start":"2023-03-01","end":"2023-04-30","color":"rgba(233,30,99,0.15)"},
    "Rally AI":             {"start":"2023-05-01","end":"2024-12-31","color":"rgba(76,175,80,0.12)"},
    "Dazi Trump":           {"start":"2025-01-20","end":"2026-04-19","color":"rgba(255,152,0,0.18)"},
}

def _add_events(fig, start, end, row=1, col=1):
    """Aggiunge bande verticali colorate per gli eventi storici."""
    for name, meta in EVENTS_META.items():
        es = pd.Timestamp(meta["start"]); ee = pd.Timestamp(meta["end"])
        s  = pd.Timestamp(start);         e  = pd.Timestamp(end)
        if ee < s or es > e:
            continue
        fig.add_vrect(
            x0=max(es,s).strftime("%Y-%m-%d"),
            x1=min(ee,e).strftime("%Y-%m-%d"),
            fillcolor=meta["color"], layer="below", line_width=0,
            annotation_text=name, annotation_position="top left",
            annotation_font_size=9,
            row=row, col=col)
    return fig

# ── Helper: carica prezzi ───────────────────────────────────────────────────
def _load(tickers, start, end):
    if isinstance(tickers, str):
        tickers = [t.strip().upper() for t in tickers.replace("\n",",").split(",") if t.strip()]
    raw = yf.download(tickers, start=start, end=end, auto_adjust=True, progress=False)
    prices = raw["Close"] if len(tickers)>1 else raw[["Close"]]
    if len(tickers)==1: prices.columns = tickers
    empty = prices.columns[prices.isna().all()].tolist()
    if empty: prices = prices.drop(columns=empty)
    prices.dropna(how="all", inplace=True); prices.ffill(inplace=True); prices.bfill(inplace=True)
    return prices

def _returns(prices, log=True):
    if log: return np.log(prices/prices.shift(1)).dropna()
    return prices.pct_change().dropna()

# ════════════════════════════════════════════════════════════════════════════
# LAYOUT
# ════════════════════════════════════════════════════════════════════════════
app = dash.Dash(
    __name__,
    external_stylesheets=[dbc.themes.FLATLY],
    suppress_callback_exceptions=True)
app.title = "Dashboard Finanziaria Quantitativa"

# ── Sidebar comune ──────────────────────────────────────────────────────────
def _sidebar():
    return dbc.Card([
        dbc.CardHeader(html.B("⚙️ Configurazione", className="text-primary")),
        dbc.CardBody([
            dbc.Label("Tickers (virgola)", className="fw-bold"),
            dbc.Textarea(id="inp-tickers", value=DEFAULT_TICKERS,
                         rows=4, style={"fontSize":"11px"}),
            dbc.Label("Benchmark", className="fw-bold mt-2"),
            dbc.Input(id="inp-benchmark", value=DEFAULT_BENCHMARK, type="text"),
            dbc.Label("Data Inizio", className="fw-bold mt-2"),
            dbc.Input(id="inp-start", value=DEFAULT_START, type="text"),
            dbc.Label("Data Fine", className="fw-bold mt-2"),
            dbc.Input(id="inp-end", value=DEFAULT_END, type="text"),
            dbc.Label("Risk-Free BCE (%)", className="fw-bold mt-2"),
            dcc.Slider(id="sl-rf", min=0, max=8, step=0.25, value=3.0,
                       marks={0:"0%",2:"2%",4:"4%",6:"6%",8:"8%"},
                       tooltip={"placement":"bottom","always_visible":True}),
            dbc.Label("Max peso (%) per asset", className="fw-bold mt-2"),
            dcc.Slider(id="sl-maxw", min=5, max=50, step=5, value=25,
                       marks={5:"5",20:"20",35:"35",50:"50"},
                       tooltip={"placement":"bottom","always_visible":True}),
            dbc.Label("Rolling window (giorni)", className="fw-bold mt-2"),
            dcc.Slider(id="sl-roll", min=60, max=500, step=10, value=250,
                       marks={60:"60",125:"125",250:"250",500:"500"},
                       tooltip={"placement":"bottom","always_visible":True}),
            dbc.Button("🚀 CARICA & ANALIZZA",
                       id="btn-load", color="primary", className="mt-3 w-100"),
            html.Div(id="load-status", className="mt-2 small text-muted"),
        ])
    ], style={"minHeight":"100vh", "fontSize":"13px"})

# ── Tabs ────────────────────────────────────────────────────────────────────
_TABS = [
    ("tab-portfolio",  "📊 Analisi Portafoglio"),
    ("tab-corr",       "🔗 Matrice Correlazioni"),
    ("tab-finanziaria","📈 Analisi Finanziaria"),
    ("tab-frontier",   "🎯 Frontiera Efficiente"),
    ("tab-style",      "🔬 Style Analysis"),
    ("tab-returns",    "📉 Rendimenti Storici"),
    ("tab-arima",      "🔮 Analisi ARIMA"),
    ("tab-rolling",    "🌊 Analisi Rolling"),
    ("tab-lstm",       "🤖 Previsione LSTM"),
    ("tab-compare",    "🏆 Confronto Portafogli"),
]

app.layout = dbc.Container([
    # ── Header ───────────────────────────────────────────────────────────
    dbc.Row([dbc.Col(html.Div([
        html.H3("📊 Dashboard Finanziaria Quantitativa", className="text-white mb-0"),
        html.Small("IS20 · Frontiera Efficiente · Style Analysis Newey-West · ARIMA · GARCH · Dazi Trump",
                   className="text-white-50")
    ], style={"background":"linear-gradient(135deg,#1F4E78,#2E86AB)",
               "padding":"12px 20px","borderRadius":"8px","marginBottom":"10px"}))]),

    # ── Body: sidebar + tabs ──────────────────────────────────────────────
    dbc.Row([
        dbc.Col(_sidebar(), width=3),
        dbc.Col([
            dcc.Tabs(id="main-tabs", value="tab-portfolio",
                     style={"fontSize":"12px"},
                     children=[dcc.Tab(label=lbl, value=val) for val,lbl in _TABS]),
            html.Div(id="tab-content", style={"padding":"10px 0"})
        ], width=9)
    ]),

    # Store dati
    dcc.Store(id="store-prices"),
    dcc.Store(id="store-returns"),
    dcc.Store(id="store-bench"),
    dcc.Store(id="store-ef"),     # efficient frontier results
    dcc.Store(id="store-style"),  # style analysis results

], fluid=True, style={"backgroundColor":"#f4f6f9","minHeight":"100vh","padding":"10px"})

# ════════════════════════════════════════════════════════════════════════════
# CALLBACK 0 — Carica dati e salva negli Store
# ════════════════════════════════════════════════════════════════════════════
@app.callback(
    Output("store-prices","data"),
    Output("store-returns","data"),
    Output("store-bench","data"),
    Output("load-status","children"),
    Input("btn-load","n_clicks"),
    State("inp-tickers","value"),
    State("inp-benchmark","value"),
    State("inp-start","value"),
    State("inp-end","value"),
    prevent_initial_call=True)
def load_data(n, tickers_str, bench, start, end):
    tickers = [t.strip().upper() for t in tickers_str.replace("\n",",").split(",") if t.strip()]
    try:
        p = _load(tickers, start, end)
        r = _returns(p)
        pb= _load([bench], start, end)
        rb= _returns(pb)
        valid = list(p.columns)
        msg = (f"✅ {len(p)} giorni · {len(valid)} titoli: {', '.join(valid[:6])}"
               f"{'...' if len(valid)>6 else ''}")
        return (p.to_json(date_format="iso"),
                r.to_json(date_format="iso"),
                pd.concat([pb, rb.rename(columns={bench:bench+"_ret"})],axis=1).to_json(date_format="iso"),
                msg)
    except Exception as e:
        return None, None, None, f"❌ {e}"

# ════════════════════════════════════════════════════════════════════════════
# CALLBACK 1 — Routing tab content
# ════════════════════════════════════════════════════════════════════════════
@app.callback(
    Output("tab-content","children"),
    Input("main-tabs","value"))
def render_tab(tab):
    tabs_map = {
        "tab-portfolio":   _layout_portfolio(),
        "tab-corr":        _layout_corr(),
        "tab-finanziaria": _layout_finanziaria(),
        "tab-frontier":    _layout_frontier(),
        "tab-style":       _layout_style(),
        "tab-returns":     _layout_returns(),
        "tab-arima":       _layout_arima(),
        "tab-rolling":     _layout_rolling(),
        "tab-lstm":        _layout_lstm(),
        "tab-compare":     _layout_compare(),
    }
    return tabs_map.get(tab, html.Div("Tab non trovato"))

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — ANALISI PORTAFOGLIO
# ════════════════════════════════════════════════════════════════════════════
def _layout_portfolio():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Label("Benchmark"),
                dcc.Dropdown(id="dd-bench-p1", options=[], value=None,
                             placeholder="Benchmark P1"),
                dbc.Label("Rolling Window (Giorni)", className="mt-2"),
                dcc.Input(id="inp-roll-p", type="number", value=250,
                          className="form-control"),
                dbc.Button("▶ Aggiorna", id="btn-port", color="success",
                           className="mt-2 w-100"),
            ], width=2),
            dbc.Col([
                dcc.Graph(id="g-cumret", style={"height":"300px"}),
                dcc.Graph(id="g-ir",     style={"height":"220px"}),
                dcc.Graph(id="g-sharpe-tev", style={"height":"220px"}),
            ], width=10)
        ]),
        dbc.Row([
            dbc.Col(html.Div(id="tbl-port-stats"), width=12)
        ], className="mt-2")
    ])

@app.callback(
    Output("dd-bench-p1","options"),
    Input("store-prices","data"))
def update_bench_options(data):
    if not data: return []
    p = pd.read_json(data, convert_dates=True)
    return [{"label":c,"value":c} for c in p.columns]

@app.callback(
    Output("g-cumret","figure"),
    Output("g-ir","figure"),
    Output("g-sharpe-tev","figure"),
    Output("tbl-port-stats","children"),
    Input("btn-port","n_clicks"),
    State("store-returns","data"),
    State("store-bench","data"),
    State("dd-bench-p1","value"),
    State("inp-roll-p","value"),
    State("sl-rf","value"),
    State("inp-start","value"),
    State("inp-end","value"),
    prevent_initial_call=True)
def update_portfolio(n, ret_data, bench_data, bench_col, roll_w, rf_pct, start, end):
    if not ret_data:
        empty = go.Figure(); empty.update_layout(template="plotly_white")
        return empty, empty, empty, ""
    r = pd.read_json(ret_data, convert_dates=True)
    rf = rf_pct/100
    roll = int(roll_w or 250)
    tickers = list(r.columns)

    # ── Calcola portafogli P1/P2/P3 ─────────────────────────────────────
    n_a = len(tickers)
    mu  = r.mean()*252; cov = r.cov()*252
    # P1 = IS20 passivo
    w1_raw = np.array([IS20_W.get(t,0.0) for t in tickers])
    w1 = w1_raw/w1_raw.sum() if w1_raw.sum()>0.05 else np.ones(n_a)/n_a
    # P2 = Max-Sharpe
    def neg_sharpe(w):
        r_ = w@mu; v_ = np.sqrt(w@cov@w); return -(r_-rf)/v_ if v_>0 else 0
    res = minimize(neg_sharpe, np.ones(n_a)/n_a, method="SLSQP",
                   bounds=[(0,.35)]*n_a,
                   constraints=[{"type":"eq","fun":lambda w:np.sum(w)-1}])
    w2 = res.x
    # P3 = Min-Var
    res3 = minimize(lambda w: w@cov@w, np.ones(n_a)/n_a, method="SLSQP",
                    bounds=[(0,.35)]*n_a,
                    constraints=[{"type":"eq","fun":lambda w:np.sum(w)-1}])
    w3 = res3.x

    port_rets = {}
    for name, w in [("P1 IS20 Passivo",w1),("P2 Max-Sharpe",w2),("P3 Min-Var",w3)]:
        port_rets[name] = pd.Series(r.values @ w, index=r.index)

    # ── Benchmark ────────────────────────────────────────────────────────
    if bench_data and bench_col:
        bd = pd.read_json(bench_data, convert_dates=True)
        if bench_col+"_ret" in bd.columns:
            port_rets[f"Benchmark ({bench_col})"] = bd[bench_col+"_ret"].reindex(r.index).fillna(0)

    colors_p = {"P1 IS20 Passivo":"#FF6B35","P2 Max-Sharpe":"#4CAF50",
                "P3 Min-Var":"#2196F3","Benchmark ("+str(bench_col)+")":"#9E9E9E"}

    # ── 1. Cumulative Returns ─────────────────────────────────────────────
    fig1 = go.Figure()
    for name, ret in port_rets.items():
        cum = np.cumsum(ret) * 100
        fig1.add_trace(go.Scatter(x=cum.index, y=cum.values, name=name,
                                   line=dict(color=colors_p.get(name,"#607D8B"), width=2)))
    _add_events(fig1, start, end)
    fig1.update_layout(title="Rendimenti Cumulativi (%)", template="plotly_white",
                       legend=dict(orientation="h",y=1.02), margin=dict(t=40,b=30),
                       hovermode="x unified")

    # ── 2. Information Ratio ─────────────────────────────────────────────
    fig2 = go.Figure()
    ref_key = list(port_rets.keys())[0]
    ref_ret = port_rets[ref_key]
    for name, ret in list(port_rets.items())[1:]:
        excess = ret - ref_ret.reindex(ret.index).fillna(0)
        ir = excess.rolling(roll).mean() / excess.rolling(roll).std() * np.sqrt(252)
        fig2.add_trace(go.Scatter(x=ir.index, y=ir.values, name=name,
                                   line=dict(color=colors_p.get(name,"#607D8B"), width=1.5)))
    fig2.add_hline(y=0, line_dash="dash", line_color="red", line_width=1)
    fig2.update_layout(title=f"Information Ratio vs {ref_key} (rolling {roll}gg)",
                       template="plotly_white", showlegend=True,
                       legend=dict(orientation="h",y=1.02), margin=dict(t=40,b=30),
                       hovermode="x unified")

    # ── 3. Sharpe Ratio & TEV ────────────────────────────────────────────
    fig3 = make_subplots(rows=1, cols=2,
                         subplot_titles=["Rolling Sharpe Ratio","Tracking Error Volatility"])
    for i, (name, ret) in enumerate(port_rets.items()):
        rm = ret.rolling(roll).mean()*252
        rv = ret.rolling(roll).std()*np.sqrt(252)
        sr = (rm - rf) / rv
        col_ = colors_p.get(name,"#607D8B")
        fig3.add_trace(go.Scatter(x=sr.index, y=sr.values, name=name,
                                   line=dict(color=col_, width=1.5)), row=1, col=1)
        if name != ref_key:
            ref = port_rets.get(ref_key, ret)
            tev = (ret - ref.reindex(ret.index).fillna(0)).rolling(roll).std()*np.sqrt(252)*100
            fig3.add_trace(go.Scatter(x=tev.index, y=tev.values, name=f"TEV {name}",
                                       line=dict(color=col_, width=1.5, dash="dot"),
                                       showlegend=False), row=1, col=2)
    fig3.add_hline(y=0, line_dash="dash", line_color="gray", row=1, col=1)
    fig3.add_hrect(y0=2, y1=4, fillcolor="rgba(76,175,80,0.1)",
                   line_width=0, row=1, col=2, annotation_text="target 2-4%")
    fig3.update_layout(template="plotly_white", margin=dict(t=40,b=30),
                       hovermode="x unified", legend=dict(orientation="h",y=1.05))

    # ── Tabella statistiche ────────────────────────────────────────────
    rows_t = []
    for name, ret in port_rets.items():
        ar = ret.mean()*252*100; av = ret.std()*np.sqrt(252)*100
        sr_ = (ar-rf*100)/av
        cum = (np.exp(np.cumsum(ret))-1).iloc[-1]*100
        cs = pd.Series(np.exp(np.cumsum(ret))); mdd = ((cs-cs.cummax())/cs.cummax()).min()*100
        cal = ar/abs(mdd) if mdd!=0 else 0
        rows_t.append({"Portafoglio":name, "Ret/Y %":f"{ar:.1f}",
                       "Vol/Y %":f"{av:.1f}", "Sharpe":f"{sr_:.3f}",
                       "Cumul. %":f"{cum:.0f}", "MaxDD %":f"{mdd:.1f}",
                       "Calmar":f"{cal:.3f}"})
    tbl = dash_table.DataTable(
        data=rows_t,
        columns=[{"name":k,"id":k} for k in rows_t[0].keys()] if rows_t else [],
        style_cell={"textAlign":"center","fontSize":"12px","padding":"6px"},
        style_header={"backgroundColor":"#1F4E78","color":"white","fontWeight":"bold"},
        style_data_conditional=[
            {"if":{"filter_query":"{Portafoglio} contains 'Max-Sharpe'"},
             "backgroundColor":"rgba(76,175,80,0.15)","fontWeight":"bold"}])
    return fig1, fig2, fig3, tbl

# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — MATRICE CORRELAZIONI
# ════════════════════════════════════════════════════════════════════════════
def _layout_corr():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Label("Periodo rolling (giorni, 0=full)"),
                dcc.Input(id="inp-corr-roll", type="number", value=0, className="form-control"),
                dbc.Button("▶ Calcola", id="btn-corr", color="success", className="mt-2 w-100"),
            ], width=2),
            dbc.Col(dcc.Graph(id="g-corr", style={"height":"600px"}), width=10)
        ])
    ])

@app.callback(
    Output("g-corr","figure"),
    Input("btn-corr","n_clicks"),
    State("store-returns","data"),
    State("inp-corr-roll","value"),
    prevent_initial_call=True)
def update_corr(n, ret_data, roll):
    if not ret_data: return go.Figure()
    r = pd.read_json(ret_data, convert_dates=True)
    if roll and int(roll) > 0:
        corr = r.tail(int(roll)).corr()
        title = f"Matrice di Correlazione — ultimi {roll} giorni"
    else:
        corr = r.corr()
        title = "Matrice di Correlazione — full sample"
    mask = np.triu(np.ones_like(corr, dtype=bool), k=1)
    corr_masked = corr.where(~mask)
    fig = px.imshow(corr_masked, text_auto=".2f", aspect="auto",
                    color_continuous_scale="RdYlGn", zmin=-1, zmax=1,
                    title=title)
    fig.update_layout(template="plotly_white",
                      coloraxis_colorbar=dict(title="ρ"),
                      margin=dict(t=50,b=20))
    return fig

# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — ANALISI FINANZIARIA (statistiche descrittive + distribuzione)
# ════════════════════════════════════════════════════════════════════════════
def _layout_finanziaria():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Label("Asset da analizzare"),
                dcc.Dropdown(id="dd-fin-asset", options=[], value=None),
                dbc.Button("▶ Analizza", id="btn-fin", color="success", className="mt-2 w-100"),
            ], width=2),
            dbc.Col([
                dcc.Graph(id="g-fin-dist", style={"height":"350px"}),
                dcc.Graph(id="g-fin-qq",   style={"height":"300px"}),
            ], width=5),
            dbc.Col([
                dcc.Graph(id="g-fin-prices", style={"height":"350px"}),
                html.Div(id="tbl-fin-stats"),
            ], width=5)
        ])
    ])

@app.callback(
    Output("dd-fin-asset","options"),
    Input("store-prices","data"))
def fin_options(data):
    if not data: return []
    p = pd.read_json(data, convert_dates=True)
    return [{"label":c,"value":c} for c in p.columns]

@app.callback(
    Output("g-fin-dist","figure"),
    Output("g-fin-qq","figure"),
    Output("g-fin-prices","figure"),
    Output("tbl-fin-stats","children"),
    Input("btn-fin","n_clicks"),
    State("store-prices","data"),
    State("store-returns","data"),
    State("dd-fin-asset","value"),
    State("inp-start","value"),
    State("inp-end","value"),
    prevent_initial_call=True)
def update_finanziaria(n, p_data, r_data, asset, start, end):
    if not p_data or not asset: return go.Figure(), go.Figure(), go.Figure(), ""
    p = pd.read_json(p_data, convert_dates=True)
    r = pd.read_json(r_data, convert_dates=True)
    if asset not in p.columns: return go.Figure(), go.Figure(), go.Figure(), ""
    ps = p[asset].dropna(); rs = r[asset].dropna()

    # Distribuzione rendimenti
    x_norm = np.linspace(rs.min(), rs.max(), 300)
    fig1 = go.Figure()
    fig1.add_trace(go.Histogram(x=rs.values, nbinsx=80, histnorm="probability density",
                                name="Distribuzione", marker_color="#2196F3", opacity=0.7))
    fig1.add_trace(go.Scatter(x=x_norm,
                               y=stats.norm.pdf(x_norm, rs.mean(), rs.std()),
                               name="Normale", line=dict(color="#F44336", width=2)))
    t_params = stats.t.fit(rs)
    fig1.add_trace(go.Scatter(x=x_norm,
                               y=stats.t.pdf(x_norm, *t_params),
                               name="Student-t", line=dict(color="#FF9800", width=2, dash="dash")))
    fig1.update_layout(title=f"Distribuzione Rendimenti — {asset}",
                       template="plotly_white", showlegend=True, margin=dict(t=40,b=20))

    # QQ-plot
    (osm, osr), (slope, intercept, _) = stats.probplot(rs, dist="norm")
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=osm, y=osr, mode="markers",
                               marker=dict(color="#9C27B0", size=4, opacity=0.6),
                               name="QQ empirico"))
    fig2.add_trace(go.Scatter(x=[osm[0],osm[-1]],
                               y=[slope*osm[0]+intercept, slope*osm[-1]+intercept],
                               mode="lines", line=dict(color="#F44336",width=2), name="Normale"))
    fig2.update_layout(title="QQ-Plot vs Normale", template="plotly_white",
                       margin=dict(t=40,b=20))

    # Prezzi
    fig3 = go.Figure()
    fig3.add_trace(go.Scatter(x=ps.index, y=ps.values, name=asset,
                               line=dict(color="#1F4E78", width=1.5)))
    ma200 = ps.rolling(200).mean()
    fig3.add_trace(go.Scatter(x=ma200.index, y=ma200.values, name="MA200",
                               line=dict(color="#FF6B35", width=1, dash="dot")))
    _add_events(fig3, start, end)
    fig3.update_layout(title=f"Prezzi — {asset}", template="plotly_white",
                       margin=dict(t=40,b=20), hovermode="x unified")

    # Statistiche
    jb_p = stats.jarque_bera(rs)[1]
    adf_p = adfuller(rs, autolag="AIC")[1]
    kpss_p = kpss(rs, regression="c", nlags="auto")[1]
    kurt = stats.kurtosis(rs); skew = stats.skew(rs)
    ann_ret = rs.mean()*252*100; ann_vol = rs.std()*np.sqrt(252)*100
    sr = (ann_ret - 3.0) / ann_vol
    cs = pd.Series(np.exp(np.cumsum(rs))); mdd = ((cs-cs.cummax())/cs.cummax()).min()*100
    data_t = [
        {"Statistica":"Rendimento annuo %","Valore":f"{ann_ret:.2f}"},
        {"Statistica":"Volatilità annua %","Valore":f"{ann_vol:.2f}"},
        {"Statistica":"Sharpe Ratio","Valore":f"{sr:.3f}"},
        {"Statistica":"Max Drawdown %","Valore":f"{mdd:.2f}"},
        {"Statistica":"Skewness","Valore":f"{skew:.3f}"},
        {"Statistica":"Excess Kurtosis","Valore":f"{kurt:.3f}"},
        {"Statistica":"Jarque-Bera p","Valore":f"{jb_p:.4f} {'✅' if jb_p>0.05 else '⚠️'}"},
        {"Statistica":"ADF p (stazionarietà)","Valore":f"{adf_p:.4f} {'✅' if adf_p<0.05 else '❌'}"},
        {"Statistica":"KPSS p","Valore":f"{kpss_p:.4f} {'✅' if kpss_p>0.05 else '❌'}"},
    ]
    tbl = dash_table.DataTable(data=data_t,
        columns=[{"name":k,"id":k} for k in data_t[0].keys()],
        style_cell={"fontSize":"12px","padding":"5px"},
        style_header={"backgroundColor":"#1F4E78","color":"white","fontWeight":"bold"})
    return fig1, fig2, fig3, tbl

# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — FRONTIERA EFFICIENTE
# ════════════════════════════════════════════════════════════════════════════
def _layout_frontier():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Label("N. simulazioni MC"),
                dcc.Input(id="inp-nsim", type="number", value=5000, className="form-control"),
                dbc.Label("Misura rischio", className="mt-2"),
                dbc.RadioItems(id="ri-risk",
                    options=[{"label":"Volatilità","value":"vol"},
                             {"label":"VaR 5%","value":"var5"},
                             {"label":"VaR 1%","value":"var1"}],
                    value="vol", inline=True),
                dbc.Checklist(id="chk-arima",
                    options=[{"label":"Modalità ARIMA attiva","value":"arima"}],
                    value=[], className="mt-2"),
                dbc.Label("Horizon ARIMA (gg)", className="mt-1"),
                dcc.Input(id="inp-arima-h", type="number", value=21, className="form-control"),
                dbc.Button("🎯 Calcola Frontiera", id="btn-frontier",
                           color="primary", className="mt-3 w-100"),
            ], width=2),
            dbc.Col([
                dcc.Graph(id="g-frontier", style={"height":"500px"}),
                dcc.Graph(id="g-frontier-cum", style={"height":"250px"}),
            ], width=7),
            dbc.Col([
                html.H6("Allocazione Max-Sharpe", className="text-primary"),
                dcc.Graph(id="g-pie-ms", style={"height":"280px"}),
                html.H6("Allocazione Min-Var", className="text-primary mt-2"),
                dcc.Graph(id="g-pie-mv", style={"height":"280px"}),
            ], width=3)
        ])
    ])

@app.callback(
    Output("g-frontier","figure"),
    Output("g-frontier-cum","figure"),
    Output("g-pie-ms","figure"),
    Output("g-pie-mv","figure"),
    Output("store-ef","data"),
    Input("btn-frontier","n_clicks"),
    State("store-returns","data"),
    State("store-prices","data"),
    State("inp-nsim","value"),
    State("sl-maxw","value"),
    State("sl-rf","value"),
    State("ri-risk","value"),
    State("chk-arima","value"),
    State("inp-arima-h","value"),
    State("inp-start","value"),
    State("inp-end","value"),
    prevent_initial_call=True)
def update_frontier(n, ret_data, p_data, n_sim, max_w_pct, rf_pct, risk_type,
                    arima_chk, arima_h, start, end):
    if not ret_data: return go.Figure(), go.Figure(), go.Figure(), go.Figure(), None
    r = pd.read_json(ret_data, convert_dates=True)
    clean = r.dropna(axis=1, how="all").dropna()
    if clean.shape[1]<2: return go.Figure(), go.Figure(), go.Figure(), go.Figure(), None

    tickers = list(clean.columns)
    n_a = len(tickers)
    mu  = clean.mean()*252
    cov = clean.cov()*252
    rf  = rf_pct/100
    mw  = max_w_pct/100
    n_sim = int(n_sim or 5000)

    # ARIMA override
    if "arima" in (arima_chk or []):
        mu_arima = {}
        for t in tickers:
            s = clean[t].dropna()
            if len(s) >= 50:
                try:
                    fc = ARIMA(s, order=(1,0,1)).fit().forecast(steps=int(arima_h or 21))
                    mu_arima[t] = float(fc.mean())*252
                except Exception:
                    mu_arima[t] = float(s.mean())*252
            else:
                mu_arima[t] = float(s.mean())*252
        mu = pd.Series(mu_arima)

    def _risk(w):
        v = np.sqrt(w@cov.values@w)
        if risk_type=="vol": return v
        port_r = clean.values@w
        if risk_type=="var5": return -np.percentile(port_r, 5)
        return -np.percentile(port_r, 1)

    # Monte Carlo
    np.random.seed(42)
    sims = {"rets":[],"vols":[],"sharpes":[],"ws":[]}
    for _ in range(n_sim):
        w = np.random.dirichlet(np.ones(n_a))
        w = np.clip(w, 0, mw); w /= w.sum()
        rv= _risk(w); re= w@mu.values; sh= (re-rf)/rv
        sims["rets"].append(re*100); sims["vols"].append(rv*100)
        sims["sharpes"].append(sh); sims["ws"].append(w)

    # Ottimizzazione
    def _opt_sharpe(w):
        rv= _risk(w); re= w@mu.values
        return -(re-rf)/rv if rv>0 else 0
    res_ms = minimize(_opt_sharpe, np.ones(n_a)/n_a, method="SLSQP",
                       bounds=[(0,mw)]*n_a,
                       constraints=[{"type":"eq","fun":lambda w:np.sum(w)-1}])
    ms_w = res_ms.x
    ms_v = _risk(ms_w)*100; ms_r = (ms_w@mu.values)*100
    ms_sh= (ms_r/100-rf)/(_risk(ms_w))

    res_mv = minimize(lambda w: w@cov.values@w, np.ones(n_a)/n_a, method="SLSQP",
                       bounds=[(0,mw)]*n_a,
                       constraints=[{"type":"eq","fun":lambda w:np.sum(w)-1}])
    mv_w = res_mv.x
    mv_v = _risk(mv_w)*100; mv_r = (mv_w@mu.values)*100

    # Frontiera
    r_min = mv_r/100; r_max = float(mu.max())
    frontier_pts = []
    for target in np.linspace(r_min, r_max, 80):
        res_f = minimize(lambda w: w@cov.values@w, np.ones(n_a)/n_a, method="SLSQP",
                          bounds=[(0,mw)]*n_a,
                          constraints=[{"type":"eq","fun":lambda w:np.sum(w)-1},
                                       {"type":"eq","fun":lambda w:w@mu.values-target}])
        if res_f.success:
            frontier_pts.append((_risk(res_f.x)*100, res_f.x@mu.values*100))

    # ── Grafico frontiera ─────────────────────────────────────────────────
    fig1 = go.Figure()
    fig1.add_trace(go.Scatter(
        x=sims["vols"], y=sims["rets"], mode="markers",
        marker=dict(color=sims["sharpes"], colorscale="RdYlGn", size=4, opacity=0.5,
                    colorbar=dict(title="Sharpe",x=1.02)),
        name="Simulazioni MC", hovertemplate="Vol:%{x:.1f}%<br>Ret:%{y:.1f}%"))
    if frontier_pts:
        fx,fy = zip(*frontier_pts)
        fig1.add_trace(go.Scatter(x=list(fx), y=list(fy), mode="lines",
                                   line=dict(color="#1F4E78",width=3), name="Frontiera"))
    fig1.add_trace(go.Scatter(x=[ms_v], y=[ms_r], mode="markers+text",
                               marker=dict(symbol="star",size=18,color="gold",
                                           line=dict(color="black",width=1.5)),
                               text=[f"Max Sharpe<br>SR:{ms_sh:.2f}"],
                               textposition="top center", name="Max Sharpe"))
    fig1.add_trace(go.Scatter(x=[mv_v], y=[mv_r], mode="markers+text",
                               marker=dict(symbol="diamond",size=14,color="cyan",
                                           line=dict(color="black",width=1.5)),
                               text=["Min Var"], textposition="top center", name="Min Var"))
    # Singoli asset
    for i, t in enumerate(tickers):
        v_i = np.sqrt(cov.iloc[i,i])*100; r_i = float(mu.iloc[i])*100
        fig1.add_trace(go.Scatter(x=[v_i], y=[r_i], mode="markers+text",
                                   marker=dict(size=8, color=COLORS_P[i%len(COLORS_P)]),
                                   text=[t], textposition="top right",
                                   name=t, showlegend=False))
    xlabel = {"vol":"Volatilità (%)","var5":"VaR 5% (%)","var1":"VaR 1% (%)"}
    title_sfx = " [ARIMA-adjusted]" if "arima" in (arima_chk or []) else " [Standard]"
    fig1.update_layout(
        title=f"Frontiera Efficiente{title_sfx} — Vincoli: 0%-{max_w_pct}%",
        xaxis_title=xlabel.get(risk_type,"Rischio (%)"),
        yaxis_title="Rendimento Atteso (%)",
        template="plotly_white", margin=dict(t=50,b=40),
        legend=dict(orientation="h",y=1.02))

    # ── Rendimenti cumulativi frontiera ───────────────────────────────────
    p2 = pd.read_json(p_data, convert_dates=True)[tickers].dropna()
    r2 = _returns(p2)
    # IS20
    w1_raw = np.array([IS20_W.get(t,0.0) for t in tickers])
    w1 = w1_raw/w1_raw.sum() if w1_raw.sum()>0.01 else np.ones(n_a)/n_a
    fig2 = go.Figure()
    for nm, w_ in [("IS20 Passivo",w1),("Max-Sharpe",ms_w),("Min-Var",mv_w)]:
        cum = np.cumsum(r2.values@w_)*100
        col_ = {"IS20 Passivo":"#FF6B35","Max-Sharpe":"#4CAF50","Min-Var":"#2196F3"}[nm]
        fig2.add_trace(go.Scatter(x=r2.index, y=cum, name=nm,
                                   line=dict(color=col_, width=2)))
    _add_events(fig2, start, end)
    fig2.update_layout(title="Rendimenti Cumulativi — F1/F2/F3",
                       template="plotly_white", margin=dict(t=40,b=20),
                       hovermode="x unified", legend=dict(orientation="h",y=1.02))

    # ── Torte ─────────────────────────────────────────────────────────────
    def _pie(weights, title):
        mask = weights > 0.005
        fig = go.Figure(go.Pie(
            labels=[tickers[i] for i in range(n_a) if mask[i]],
            values=weights[mask], hole=0.35,
            marker_colors=COLORS_P[:mask.sum()],
            textinfo="label+percent"))
        fig.update_layout(title=title, template="plotly_white",
                          margin=dict(t=40,b=10,l=10,r=10),
                          showlegend=False)
        return fig

    ef_data = {"ms_weights":ms_w.tolist(),"mv_weights":mv_w.tolist(),
               "tickers":tickers,"ms_sharpe":float(ms_sh)}
    return fig1, fig2, _pie(ms_w, f"Max-Sharpe SR:{ms_sh:.2f}"), _pie(mv_w, "Min-Var"), ef_data

# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — STYLE ANALYSIS (Newey-West HAC)
# ════════════════════════════════════════════════════════════════════════════
def _layout_style():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Label("Asset Y"),
                dcc.Dropdown(id="dd-style-y", options=[], value=None),
                dbc.Label("Fattori di stile X (ETF)", className="mt-2"),
                dcc.Dropdown(id="dd-style-x", options=[], value=[], multi=True),
                dbc.Label("Std Error", className="mt-2"),
                dbc.RadioItems(id="ri-se",
                    options=[{"label":"OLS","value":"ols"},
                             {"label":"HC3","value":"HC3"},
                             {"label":"HAC (Newey-West)","value":"HAC"}],
                    value="HAC", inline=False),
                dbc.Label("Finestra rolling (mesi)", className="mt-2"),
                dcc.Input(id="inp-style-roll", type="number", value=36, className="form-control"),
                dbc.Button("▶ Esegui Style Analysis", id="btn-style",
                           color="primary", className="mt-3 w-100"),
            ], width=3),
            dbc.Col([
                html.Div(id="div-style-stats"),
                dcc.Graph(id="g-style-rolling", style={"height":"300px"}),
            ], width=9)
        ])
    ])

@app.callback(
    Output("dd-style-y","options"), Output("dd-style-x","options"),
    Input("store-returns","data"))
def style_options(data):
    if not data: return [],[]
    r = pd.read_json(data, convert_dates=True)
    opts = [{"label":c,"value":c} for c in r.columns]
    return opts, opts

@app.callback(
    Output("div-style-stats","children"),
    Output("g-style-rolling","figure"),
    Output("store-style","data"),
    Input("btn-style","n_clicks"),
    State("store-returns","data"),
    State("dd-style-y","value"),
    State("dd-style-x","value"),
    State("ri-se","value"),
    State("inp-style-roll","value"),
    prevent_initial_call=True)
def update_style(n, ret_data, asset_y, assets_x, se_type, roll_m):
    if not ret_data or not asset_y or not assets_x:
        return dbc.Alert("Seleziona Asset Y e almeno un fattore X.",color="warning"), go.Figure(), None
    r = pd.read_json(ret_data, convert_dates=True)
    if asset_y not in r.columns:
        return dbc.Alert("Asset Y non trovato.",color="danger"), go.Figure(), None
    valid_x = [x for x in assets_x if x in r.columns]
    if not valid_x:
        return dbc.Alert("Nessun fattore X trovato.",color="danger"), go.Figure(), None

    combined = r[[asset_y]+valid_x].dropna()
    y = combined[asset_y]
    X = add_constant(combined[valid_x])

    model = OLS(y, X)
    if se_type=="HAC":
        result = model.fit(cov_type="HAC", cov_kwds={"maxlags":5})
    elif se_type=="HC3":
        result = model.fit(cov_type="HC3")
    else:
        result = model.fit()

    resid = result.resid
    dw_val = durbin_watson(resid)
    jb_stat, jb_p = stats.jarque_bera(resid)[:2]
    lb_p = acorr_ljungbox(resid, lags=[10], return_df=True)["lb_pvalue"].values[0]
    try: bp_stat, bp_p, _, _ = het_breuschpagan(resid, X)
    except: bp_stat, bp_p = np.nan, np.nan

    # ── Header statistiche ────────────────────────────────────────────────
    def _badge(val, ok_cond, ok_txt, fail_txt):
        return dbc.Badge(ok_txt if ok_cond else fail_txt,
                         color="success" if ok_cond else "danger",
                         className="ms-1")

    stats_card = dbc.Card([
        dbc.CardHeader(html.B(f"Style Analysis: {asset_y} | Std Error: {se_type} | {len(combined)} osservazioni")),
        dbc.CardBody([
            dbc.Row([
                dbc.Col([
                    html.B(f"R² = {result.rsquared*100:.2f}%"),
                    html.Span(f"  |  Adj-R² = {result.rsquared_adj*100:.2f}%"),
                    html.Br(),
                    html.Span(f"F-stat = {result.fvalue:.2f}  (p = {result.f_pvalue:.2e})  "),
                    html.Br(),
                    html.Span(f"AIC = {result.aic:.2f}  |  BIC = {result.bic:.2f}"),
                ], width=4),
                dbc.Col([
                    html.Span("Durbin-Watson: "),
                    html.B(f"{dw_val:.4f}"),
                    _badge(dw_val, 1.5<dw_val<2.5, " ≈ 2 (no autocorr.)", " ≠ 2 (autocorr.!)"),
                    html.Br(),
                    html.Span("Jarque-Bera p: "),
                    html.B(f"{jb_p:.4f}"),
                    _badge(jb_p, jb_p>0.05, " Normale", " Non normale"),
                    html.Br(),
                    html.Span("Ljung-Box p(10): "),
                    html.B(f"{lb_p:.4f}"),
                    _badge(lb_p, lb_p>0.05, " No AC", " AC!"),
                ], width=4),
                dbc.Col([
                    html.Span("Breusch-Pagan p: "),
                    html.B(f"{bp_p:.4f}" if not np.isnan(bp_p) else "n.d."),
                    _badge(bp_p, not np.isnan(bp_p) and bp_p>0.05, " Omoscedast.", " Eteroscedast."),
                    html.Br(),
                    html.Span(f"Skewness: {stats.skew(resid):.3f}  |  Kurtosis: {stats.kurtosis(resid):.3f}"),
                ], width=4),
            ]),
        ])
    ], className="mb-2")

    # ── Tabella coefficienti ─────────────────────────────────────────────
    coef_data = []
    for var in result.params.index:
        p_val = result.pvalues[var]
        sig = "***" if p_val<0.001 else "**" if p_val<0.01 else "*" if p_val<0.05 else "." if p_val<0.1 else ""
        coef_data.append({
            "Variabile":var,
            "Coeff.":f"{result.params[var]:.6f}",
            "Std Err":f"{result.bse[var]:.6f}",
            "t-stat":f"{result.tvalues[var]:.4f}",
            "p-val":f"{p_val:.4e}",
            "Sig.":sig,
            "IC95 inf":f"{result.conf_int().loc[var,0]:.6f}",
            "IC95 sup":f"{result.conf_int().loc[var,1]:.6f}",
        })
    tbl_coef = dash_table.DataTable(
        data=coef_data,
        columns=[{"name":k,"id":k} for k in coef_data[0].keys()] if coef_data else [],
        style_cell={"fontSize":"11px","padding":"4px","textAlign":"right"},
        style_header={"backgroundColor":"#1F4E78","color":"white","fontWeight":"bold"},
        style_data_conditional=[
            {"if":{"filter_query":'{Sig.} contains "*"'},
             "fontWeight":"bold","backgroundColor":"rgba(76,175,80,0.1)"},
            {"if":{"column_id":"Variabile"},"textAlign":"left"},
        ])

    # ── Rolling Style Weights ─────────────────────────────────────────────
    roll_days = int(roll_m or 36) * 21  # mesi → giorni approssimativi
    fig = go.Figure()
    if len(combined) > roll_days + 10:
        dates_r = combined.index[roll_days:]
        weights_ts = {v: [] for v in valid_x}
        for i in range(roll_days, len(combined)):
            window = combined.iloc[i-roll_days:i]
            y_w = window[asset_y]; X_w = add_constant(window[valid_x])
            try:
                res_w = OLS(y_w, X_w).fit()
                coefs = np.maximum(res_w.params[valid_x].values, 0)
                total = coefs.sum()
                coefs = coefs/total if total>0 else np.ones(len(valid_x))/len(valid_x)
            except Exception:
                coefs = np.ones(len(valid_x))/len(valid_x)
            for j, v in enumerate(valid_x):
                weights_ts[v].append(coefs[j]*100)

        for i, v in enumerate(valid_x):
            fig.add_trace(go.Bar(x=dates_r, y=weights_ts[v], name=v,
                                  marker_color=COLORS_P[i%len(COLORS_P)]))
        fig.update_layout(barmode="stack",
                          title=f"Style Weights Rolling — finestra {roll_m} mesi",
                          yaxis_title="Peso (%)", template="plotly_white",
                          legend=dict(orientation="h",y=1.02),
                          margin=dict(t=50,b=20))
    else:
        fig.update_layout(title="Dati insufficienti per rolling",template="plotly_white")

    style_res = {"r2":result.rsquared,"dw":dw_val,"jb_p":jb_p}
    return [stats_card, tbl_coef], fig, style_res

# ════════════════════════════════════════════════════════════════════════════
# TAB 6 — RENDIMENTI STORICI
# ════════════════════════════════════════════════════════════════════════════
def _layout_returns():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Label("Asset da visualizzare"),
                dcc.Dropdown(id="dd-ret-assets", options=[], value=[], multi=True),
                dbc.Label("Tipo rendimento", className="mt-2"),
                dbc.RadioItems(id="ri-ret-type",
                    options=[{"label":"Log cumulativo","value":"logcum"},
                             {"label":"Cumulativo €","value":"euro"},
                             {"label":"Annuale a barre","value":"annual"},
                             {"label":"Heatmap mensile","value":"heat"}],
                    value="logcum"),
                dbc.Button("▶ Mostra", id="btn-ret", color="success", className="mt-2 w-100"),
            ], width=2),
            dbc.Col(dcc.Graph(id="g-returns", style={"height":"650px"}), width=10)
        ])
    ])

@app.callback(
    Output("dd-ret-assets","options"),
    Input("store-prices","data"))
def ret_options(data):
    if not data: return []
    p = pd.read_json(data, convert_dates=True)
    return [{"label":c,"value":c} for c in p.columns]

@app.callback(
    Output("g-returns","figure"),
    Input("btn-ret","n_clicks"),
    State("store-prices","data"),
    State("store-returns","data"),
    State("dd-ret-assets","value"),
    State("ri-ret-type","value"),
    State("inp-start","value"),
    State("inp-end","value"),
    prevent_initial_call=True)
def update_returns(n, p_data, r_data, assets, ret_type, start, end):
    if not r_data or not assets: return go.Figure()
    r = pd.read_json(r_data, convert_dates=True)
    sel = [a for a in assets if a in r.columns]
    if not sel: return go.Figure()
    r_sel = r[sel]

    if ret_type == "logcum":
        fig = go.Figure()
        for i, a in enumerate(sel):
            cum = np.cumsum(r_sel[a])*100
            fig.add_trace(go.Scatter(x=cum.index, y=cum.values, name=a,
                                      line=dict(color=COLORS_P[i%len(COLORS_P)], width=2)))
        _add_events(fig, start, end)
        fig.update_layout(title="Rendimenti Log Cumulativi (%)",
                          template="plotly_white", hovermode="x unified",
                          legend=dict(orientation="h",y=1.02))

    elif ret_type == "euro":
        p = pd.read_json(p_data, convert_dates=True)[sel]
        fig = go.Figure()
        for i, a in enumerate(sel):
            ps = p[a].dropna()
            norm = ps/ps.iloc[0]*100
            fig.add_trace(go.Scatter(x=norm.index, y=norm.values, name=a,
                                      line=dict(color=COLORS_P[i%len(COLORS_P)], width=2)))
        _add_events(fig, start, end)
        fig.update_layout(title="Crescita di €100 investiti",
                          template="plotly_white", hovermode="x unified",
                          legend=dict(orientation="h",y=1.02))

    elif ret_type == "annual":
        ann = r_sel.resample("YE").sum()*100
        fig = go.Figure()
        for i, a in enumerate(sel):
            fig.add_trace(go.Bar(x=ann.index.year, y=ann[a].values, name=a,
                                  marker_color=COLORS_P[i%len(COLORS_P)]))
        fig.add_hline(y=0, line_dash="dash", line_color="black", line_width=1)
        fig.update_layout(title="Rendimenti Annuali (%)", barmode="group",
                          template="plotly_white")

    else:  # heatmap
        asset = sel[0]
        monthly = r_sel[asset].resample("ME").sum()*100
        monthly_df = monthly.to_frame()
        monthly_df["year"]  = monthly_df.index.year
        monthly_df["month"] = monthly_df.index.month
        pivot = monthly_df.pivot(index="year", columns="month", values=asset)
        pivot.columns = ["Gen","Feb","Mar","Apr","Mag","Giu",
                         "Lug","Ago","Set","Ott","Nov","Dic"][:len(pivot.columns)]
        fig = px.imshow(pivot, text_auto=".1f", aspect="auto",
                        color_continuous_scale="RdYlGn", zmin=-15, zmax=15,
                        title=f"Heatmap Rendimenti Mensili — {asset} (%)")
        fig.update_layout(template="plotly_white")

    return fig

# ════════════════════════════════════════════════════════════════════════════
# TAB 7 — ANALISI ARIMA
# ════════════════════════════════════════════════════════════════════════════
def _layout_arima():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Label("Asset"),
                dcc.Dropdown(id="dd-arima-asset", options=[], value=None),
                dbc.Label("Ordine p", className="mt-2"),
                dcc.Input(id="inp-ap", type="number", value=1, min=0, max=5, className="form-control"),
                dbc.Label("Ordine d"),
                dcc.Input(id="inp-ad", type="number", value=0, min=0, max=2, className="form-control"),
                dbc.Label("Ordine q"),
                dcc.Input(id="inp-aq", type="number", value=1, min=0, max=5, className="form-control"),
                dbc.Label("Horizon (giorni)", className="mt-2"),
                dcc.Input(id="inp-ah", type="number", value=21, className="form-control"),
                dbc.Button("🔮 Stima ARIMA", id="btn-arima", color="primary", className="mt-3 w-100"),
            ], width=2),
            dbc.Col([
                dcc.Graph(id="g-arima-forecast", style={"height":"400px"}),
                dcc.Graph(id="g-arima-resid",    style={"height":"250px"}),
            ], width=7),
            dbc.Col([
                html.Div(id="div-arima-stats"),
            ], width=3)
        ])
    ])

@app.callback(
    Output("dd-arima-asset","options"),
    Input("store-returns","data"))
def arima_opts(data):
    if not data: return []
    r = pd.read_json(data, convert_dates=True)
    return [{"label":c,"value":c} for c in r.columns]

@app.callback(
    Output("g-arima-forecast","figure"),
    Output("g-arima-resid","figure"),
    Output("div-arima-stats","children"),
    Input("btn-arima","n_clicks"),
    State("store-returns","data"),
    State("dd-arima-asset","value"),
    State("inp-ap","value"), State("inp-ad","value"), State("inp-aq","value"),
    State("inp-ah","value"),
    prevent_initial_call=True)
def update_arima(n, ret_data, asset, ap, ad, aq, ah):
    if not ret_data or not asset: return go.Figure(), go.Figure(), ""
    r = pd.read_json(ret_data, convert_dates=True)
    if asset not in r.columns: return go.Figure(), go.Figure(), ""
    series = r[asset].dropna()
    if len(series)<50: return go.Figure(), go.Figure(), dbc.Alert("Dati insufficienti.",color="warning")

    order = (int(ap or 1), int(ad or 0), int(aq or 1))
    try:
        model_fit = ARIMA(series, order=order).fit()
    except Exception as e:
        return go.Figure(), go.Figure(), dbc.Alert(f"ARIMA errore: {e}", color="danger")

    h = int(ah or 21)
    fc_obj = model_fit.get_forecast(steps=h)
    fc_mean = fc_obj.predicted_mean
    fc_ci   = fc_obj.conf_int(alpha=0.05)

    # Indici futuri
    last = series.index[-1]
    try:
        freq = pd.infer_freq(series.index) or "B"
        fut_idx = pd.date_range(start=last, periods=h+1, freq=freq)[1:]
    except Exception:
        fut_idx = pd.RangeIndex(start=len(series), stop=len(series)+h)

    # Forecast plot
    fig1 = go.Figure()
    fig1.add_trace(go.Scatter(x=series.index[-252:], y=series.values[-252:],
                               name="Storico (1Y)", line=dict(color="#2196F3",width=1.5)))
    fig1.add_trace(go.Scatter(x=list(fut_idx), y=list(fc_mean.values),
                               name="Forecast", line=dict(color="#FF6B35",width=2,dash="dash")))
    fig1.add_trace(go.Scatter(
        x=list(fut_idx)+list(fut_idx)[::-1],
        y=list(fc_ci.iloc[:,0].values)+list(fc_ci.iloc[:,1].values)[::-1],
        fill="toself", fillcolor="rgba(255,107,53,0.15)", line_color="rgba(255,255,255,0)",
        name="IC 95%"))
    mu_ann = float(fc_mean.mean())*252*100
    fig1.update_layout(title=f"ARIMA{order} Forecast {asset} | μ_arima: {mu_ann:+.2f}%/anno",
                       template="plotly_white", margin=dict(t=50,b=20),
                       legend=dict(orientation="h",y=1.02), hovermode="x unified")

    # Residui
    resid = model_fit.resid
    fig2 = make_subplots(rows=1, cols=2, subplot_titles=["Residui nel Tempo","ACF Residui"])
    fig2.add_trace(go.Scatter(x=resid.index, y=resid.values, mode="lines",
                               line=dict(color="#9C27B0",width=0.8), name="Residui"), row=1,col=1)
    fig2.add_hline(y=0, line_dash="dash", line_color="gray", row=1, col=1)
    acf_v = sm_acf(resid, nlags=20, fft=True)[1:]
    conf = 1.96/np.sqrt(len(resid))
    fig2.add_trace(go.Bar(x=list(range(1,21)), y=acf_v.tolist(),
                           marker_color=["#F44336" if abs(v)>conf else "#4CAF50" for v in acf_v],
                           name="ACF"), row=1,col=2)
    fig2.add_hline(y=conf, line_dash="dot", line_color="red", row=1, col=2)
    fig2.add_hline(y=-conf, line_dash="dot", line_color="red", row=1, col=2)
    fig2.update_layout(template="plotly_white", margin=dict(t=30,b=20), showlegend=False)

    # Statistiche
    jb_p = stats.jarque_bera(resid)[1]
    lb_p = acorr_ljungbox(resid, lags=[10], return_df=True)["lb_pvalue"].values[0]
    data_st = [
        {"Stat":"AIC","Val":f"{model_fit.aic:.2f}"},
        {"Stat":"BIC","Val":f"{model_fit.bic:.2f}"},
        {"Stat":"Log-lik.","Val":f"{model_fit.llf:.2f}"},
        {"Stat":"σ residui","Val":f"{resid.std():.6f}"},
        {"Stat":"JB p-val","Val":f"{jb_p:.4f} {'✅' if jb_p>0.05 else '⚠️'}"},
        {"Stat":"LB p(10)","Val":f"{lb_p:.4f} {'✅' if lb_p>0.05 else '⚠️'}"},
        {"Stat":"μ ARIMA/anno","Val":f"{mu_ann:+.2f}%"},
        {"Stat":"σ ARIMA/anno","Val":f"{series.tail(63).std()*np.sqrt(252)*100:.2f}%"},
    ]
    tbl = dash_table.DataTable(data=data_st,
        columns=[{"name":k,"id":k} for k in data_st[0].keys()],
        style_cell={"fontSize":"11px","padding":"5px"},
        style_header={"backgroundColor":"#1F4E78","color":"white","fontWeight":"bold"})
    return fig1, fig2, tbl

# ════════════════════════════════════════════════════════════════════════════
# TAB 8 — ANALISI ROLLING (Volatilità + GARCH + Regime)
# ════════════════════════════════════════════════════════════════════════════
def _layout_rolling():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Label("Asset"),
                dcc.Dropdown(id="dd-roll-asset", options=[], value=None),
                dbc.Label("Finestra rolling (gg)", className="mt-2"),
                dcc.Input(id="inp-roll-w", type="number", value=21, className="form-control"),
                dbc.Label("GARCH p"),
                dcc.Input(id="inp-gp", type="number", value=1, min=1, max=3, className="form-control"),
                dbc.Label("GARCH q"),
                dcc.Input(id="inp-gq", type="number", value=1, min=1, max=3, className="form-control"),
                dbc.Button("🌊 Calcola", id="btn-roll", color="primary", className="mt-3 w-100"),
            ], width=2),
            dbc.Col([
                dcc.Graph(id="g-roll-vol",   style={"height":"250px"}),
                dcc.Graph(id="g-roll-sharpe",style={"height":"200px"}),
                dcc.Graph(id="g-garch",      style={"height":"250px"}),
                dcc.Graph(id="g-regime",     style={"height":"180px"}),
            ], width=10)
        ])
    ])

@app.callback(
    Output("dd-roll-asset","options"),
    Input("store-returns","data"),
    Input("main-tabs","value"))
def roll_opts(data, tab):
    if not data: return []
    r = pd.read_json(data, convert_dates=True)
    return [{"label":c,"value":c} for c in r.columns]

@app.callback(
    Output("g-roll-vol","figure"),
    Output("g-roll-sharpe","figure"),
    Output("g-garch","figure"),
    Output("g-regime","figure"),
    Input("btn-roll","n_clicks"),
    State("store-returns","data"),
    State("dd-roll-asset","value"),
    State("inp-roll-w","value"),
    State("inp-gp","value"),
    State("inp-gq","value"),
    State("sl-rf","value"),
    State("inp-start","value"),
    State("inp-end","value"),
    prevent_initial_call=True)
def update_rolling(n, ret_data, asset, window, gp, gq, rf_pct, start, end):
    empty = go.Figure()
    if not ret_data or not asset: return empty, empty, empty, empty
    r = pd.read_json(ret_data, convert_dates=True)
    if asset not in r.columns: return empty,empty,empty,empty
    rs = r[asset].dropna()
    w = int(window or 21); rf = (rf_pct if rf_pct is not None else 3.0)/100

    rv = rs.rolling(w).std()*np.sqrt(252)*100
    rm = rs.rolling(w).mean()*252*100
    rsh= (rm - rf*100) / rv

    # Vol
    fig1 = go.Figure()
    fig1.add_trace(go.Scatter(x=rv.index, y=rv.values, name=f"Vol rolling {w}gg",
                               fill="tozeroy", line=dict(color="#FF6B35",width=1.5)))
    _add_events(fig1, start, end)
    fig1.update_layout(title=f"Volatilità Rolling {w}gg — {asset} (% ann.)",
                       template="plotly_white", margin=dict(t=40,b=10), hovermode="x unified")

    # Sharpe
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=rsh.index, y=rsh.values, name="Sharpe rolling",
                               line=dict(color="#9C27B0",width=1.5)))
    fig2.add_hline(y=1, line_dash="dash", line_color="green", line_width=1,
                   annotation_text="SR=1")
    fig2.add_hline(y=0, line_dash="dash", line_color="red", line_width=0.8)
    fig2.update_layout(title=f"Sharpe Ratio Rolling — {asset}",
                       template="plotly_white", margin=dict(t=40,b=10), hovermode="x unified")

    # GARCH
    fig3 = go.Figure()
    try:
        am = arch_model(rs*100, p=int(gp or 1), q=int(gq or 1),
                        mean="Constant", vol="GARCH", dist="Normal")
        res_g = am.fit(disp="off")
        cv = res_g.conditional_volatility*np.sqrt(252)
        fig3.add_trace(go.Scatter(x=rv.index, y=rv.values, name=f"Rolling {w}gg",
                                   line=dict(color="#2196F3",width=1,opacity=0.7)))
        fig3.add_trace(go.Scatter(x=cv.index, y=cv.values, name=f"GARCH({gp},{gq})",
                                   line=dict(color="#FF6B35",width=2)))
        _add_events(fig3, start, end)
        fig3.update_layout(title=f"GARCH({gp},{gq}) vs Rolling — {asset}",
                           template="plotly_white", margin=dict(t=40,b=10),
                           legend=dict(orientation="h",y=1.02), hovermode="x unified")
    except Exception as e:
        fig3.update_layout(title=f"GARCH non disponibile: {e}", template="plotly_white")

    # Regime
    fig4 = go.Figure()
    q33 = rv.quantile(0.33); q66 = rv.quantile(0.66)
    _rgba = {"#4CAF50":"rgba(76,175,80,0.6)","#FF9800":"rgba(255,152,0,0.6)","#F44336":"rgba(244,67,54,0.6)"}
    for lbl, lo, hi, col in [("Bassa",-np.inf,q33,"#4CAF50"),
                               ("Media",q33,q66,"#FF9800"),
                               ("Alta",q66,np.inf,"#F44336")]:
        mask = (rv>=lo)&(rv<hi)
        y_seg = rv.where(mask)
        fig4.add_trace(go.Scatter(x=rv.index, y=y_seg.values,
                                   fill="tozeroy", name=lbl,
                                   line=dict(color=col,width=0),
                                   fillcolor=_rgba[col]))
    fig4.update_layout(title=f"Regime Volatilità — {asset}",
                       template="plotly_white", margin=dict(t=40,b=10),
                       legend=dict(orientation="h",y=1.02))
    return fig1, fig2, fig3, fig4

# ════════════════════════════════════════════════════════════════════════════
# TAB 9 — PREVISIONE LSTM (placeholder con spiegazione)
# ════════════════════════════════════════════════════════════════════════════
def _layout_lstm():
    return dbc.Card([
        dbc.CardBody([
            html.H4("🤖 Previsione LSTM", className="text-primary"),
            dbc.Alert([
                html.H5("Prossimamente disponibile"),
                html.P("Il modulo LSTM (Long Short-Term Memory) per la previsione dei prezzi "
                       "richiede TensorFlow/Keras che non è incluso nell'installazione base "
                       "per compatibilità con Colab. Per abilitarlo:"),
                dbc.Button("Installa TensorFlow", id="btn-lstm-install", color="warning",
                           size="sm", className="mb-2"),
                html.Div(id="lstm-install-status"),
                html.Hr(),
                html.P("Architettura prevista:"),
                html.Ul([
                    html.Li("Input: finestra scorrevole 60 giorni (prezzi normalizzati)"),
                    html.Li("Strati: LSTM(64) → Dropout(0.2) → LSTM(32) → Dense(1)"),
                    html.Li("Loss: MSE | Optimizer: Adam | Epochs: 50"),
                    html.Li("Output: previsione H giorni con intervallo di confidenza (MC Dropout)"),
                ])
            ], color="info"),
        ])
    ])

@app.callback(
    Output("lstm-install-status","children"),
    Input("btn-lstm-install","n_clicks"),
    prevent_initial_call=True)
def install_lstm(n):
    try:
        import subprocess, sys
        subprocess.check_call([sys.executable,"-m","pip","install","-q","tensorflow"])
        return dbc.Alert("✅ TensorFlow installato! Riavvia il notebook.", color="success")
    except Exception as e:
        return dbc.Alert(f"❌ Errore: {e}", color="danger")

# ════════════════════════════════════════════════════════════════════════════
# TAB 10 — CONFRONTO PORTAFOGLI (IS20 vs P2 vs P3)
# ════════════════════════════════════════════════════════════════════════════
def _layout_compare():
    return html.Div([
        dbc.Row([
            dbc.Col([
                dbc.Label("Rolling window (gg)"),
                dcc.Input(id="inp-cmp-roll", type="number", value=250, className="form-control"),
                dbc.Button("▶ Confronta", id="btn-cmp", color="primary", className="mt-2 w-100"),
            ], width=2),
            dbc.Col([
                dcc.Graph(id="g-cmp-cum",   style={"height":"280px"}),
                dcc.Graph(id="g-cmp-dd",    style={"height":"200px"}),
                dcc.Graph(id="g-cmp-alpha", style={"height":"200px"}),
                dcc.Graph(id="g-cmp-tev",   style={"height":"180px"}),
                html.Div(id="tbl-cmp-stats", className="mt-2"),
            ], width=10)
        ])
    ])

@app.callback(
    Output("g-cmp-cum","figure"),
    Output("g-cmp-dd","figure"),
    Output("g-cmp-alpha","figure"),
    Output("g-cmp-tev","figure"),
    Output("tbl-cmp-stats","children"),
    Input("btn-cmp","n_clicks"),
    State("store-returns","data"),
    State("store-ef","data"),
    State("inp-cmp-roll","value"),
    State("sl-rf","value"),
    State("inp-start","value"),
    State("inp-end","value"),
    prevent_initial_call=True)
def update_compare(n, ret_data, ef_data, roll, rf_pct, start, end):
    empty = go.Figure()
    if not ret_data: return empty,empty,empty,empty,""
    r = pd.read_json(ret_data, convert_dates=True)
    rf = rf_pct/100; roll = int(roll or 250)
    tickers = list(r.columns); n_a = len(tickers)

    if ef_data:
        ms_w = np.array(ef_data["ms_weights"])
        mv_w = np.array(ef_data["mv_weights"])
        ef_tickers = ef_data["tickers"]
        ms_w_full = np.array([ms_w[ef_tickers.index(t)] if t in ef_tickers else 0.0 for t in tickers])
        mv_w_full = np.array([mv_w[ef_tickers.index(t)] if t in ef_tickers else 0.0 for t in tickers])
        if ms_w_full.sum()>0: ms_w_full /= ms_w_full.sum()
        if mv_w_full.sum()>0: mv_w_full /= mv_w_full.sum()
    else:
        cov = r.cov()*252; mu = r.mean()*252
        res2 = minimize(lambda w: -(w@mu-(rf))/np.sqrt(w@cov@w),
                        np.ones(n_a)/n_a, method="SLSQP",
                        bounds=[(0,.35)]*n_a,
                        constraints=[{"type":"eq","fun":lambda w:np.sum(w)-1}])
        ms_w_full = res2.x
        res3 = minimize(lambda w: w@cov@w, np.ones(n_a)/n_a, method="SLSQP",
                        bounds=[(0,.35)]*n_a,
                        constraints=[{"type":"eq","fun":lambda w:np.sum(w)-1}])
        mv_w_full = res3.x

    w1_raw = np.array([IS20_W.get(t,0.0) for t in tickers])
    w1 = w1_raw/w1_raw.sum() if w1_raw.sum()>0.01 else np.ones(n_a)/n_a

    port_rets = {
        "IS20 Passivo (P1)":  pd.Series(r.values@w1,        index=r.index),
        "Max-Sharpe (P2)":    pd.Series(r.values@ms_w_full,  index=r.index),
        "Min-Var (P3)":       pd.Series(r.values@mv_w_full,  index=r.index),
    }
    colors_cmp = {"IS20 Passivo (P1)":"#FF6B35","Max-Sharpe (P2)":"#4CAF50","Min-Var (P3)":"#2196F3"}

    # Cumulative
    fig1 = go.Figure()
    for nm, ret in port_rets.items():
        fig1.add_trace(go.Scatter(x=ret.index, y=np.cumsum(ret)*100, name=nm,
                                   line=dict(color=colors_cmp[nm],width=2)))
    _add_events(fig1, start, end)
    fig1.update_layout(title="Performance Cumulativa (%)", template="plotly_white",
                       legend=dict(orientation="h",y=1.02), hovermode="x unified",
                       margin=dict(t=40,b=10))

    # Drawdown
    fig2 = go.Figure()
    for nm, ret in port_rets.items():
        cs = pd.Series(np.exp(np.cumsum(ret)))
        dd = (cs - cs.cummax())/cs.cummax()*100
        fig2.add_trace(go.Scatter(x=ret.index, y=dd.values, name=nm,
                                   line=dict(color=colors_cmp[nm],width=1.5)))
    fig2.add_hline(y=0, line_dash="dash", line_color="gray", line_width=0.8)
    fig2.update_layout(title="Drawdown dal Massimo (%)", template="plotly_white",
                       legend=dict(orientation="h",y=1.02), hovermode="x unified",
                       margin=dict(t=40,b=10))

    # Rolling Alpha vs P1
    p1 = port_rets["IS20 Passivo (P1)"]
    fig3 = go.Figure()
    for nm, ret in list(port_rets.items())[1:]:
        alpha = (ret - p1).rolling(roll).mean()*252*100
        fig3.add_trace(go.Scatter(x=alpha.index, y=alpha.values, name=f"α {nm} vs P1",
                                   line=dict(color=colors_cmp[nm],width=1.5)))
        fig3.add_trace(go.Scatter(
            x=alpha.index.tolist()+alpha.index.tolist()[::-1],
            y=np.where(alpha>0,alpha,0).tolist()+np.zeros(len(alpha)).tolist()[::-1],
            fill="toself", fillcolor=f"rgba(76,175,80,0.1)",
            line_color="rgba(0,0,0,0)", showlegend=False))
    fig3.add_hline(y=0, line_dash="dash", line_color="red", line_width=1)
    fig3.update_layout(title=f"Rolling Alpha vs IS20 Passivo ({roll}gg, %ann.)",
                       template="plotly_white", legend=dict(orientation="h",y=1.02),
                       hovermode="x unified", margin=dict(t=40,b=10))

    # TEV
    fig4 = go.Figure()
    for nm, ret in list(port_rets.items())[1:]:
        tev = (ret - p1).rolling(roll).std()*np.sqrt(252)*100
        fig4.add_trace(go.Scatter(x=tev.index, y=tev.values, name=f"TEV {nm}",
                                   line=dict(color=colors_cmp[nm],width=1.5)))
    fig4.add_hrect(y0=2, y1=4, fillcolor="rgba(76,175,80,0.1)", line_width=0,
                   annotation_text="Target 2-4%", annotation_position="top right")
    fig4.update_layout(title=f"Tracking Error Volatility — Target 2-4% ({roll}gg)",
                       template="plotly_white", legend=dict(orientation="h",y=1.02),
                       margin=dict(t=40,b=10))

    # Tabella
    rows_t = []
    for nm, ret in port_rets.items():
        ar = ret.mean()*252*100; av = ret.std()*np.sqrt(252)*100
        sr_ = (ar-rf*100)/av
        cum = (np.exp(np.cumsum(ret))-1).iloc[-1]*100
        cs2 = pd.Series(np.exp(np.cumsum(ret)))
        mdd = ((cs2-cs2.cummax())/cs2.cummax()).min()*100
        cal = ar/abs(mdd) if mdd!=0 else 0
        rows_t.append({"Portafoglio":nm,"Ret/Y %":f"{ar:.1f}",
                       "Vol/Y %":f"{av:.1f}","Sharpe":f"{sr_:.3f}",
                       "Cumulativo %":f"{cum:.0f}","MaxDD %":f"{mdd:.1f}",
                       "Calmar":f"{cal:.3f}"})
    tbl = dash_table.DataTable(data=rows_t,
        columns=[{"name":k,"id":k} for k in rows_t[0].keys()],
        style_cell={"textAlign":"center","fontSize":"12px","padding":"6px"},
        style_header={"backgroundColor":"#1F4E78","color":"white","fontWeight":"bold"},
        style_data_conditional=[
            {"if":{"filter_query":"{Portafoglio} contains 'Max-Sharpe'"},
             "backgroundColor":"rgba(76,175,80,0.15)","fontWeight":"bold"},
            {"if":{"filter_query":"{Portafoglio} contains 'IS20'"},
             "backgroundColor":"rgba(255,107,53,0.1)"},
        ])
    return fig1, fig2, fig3, fig4, tbl


# ── Entry point per Render.com (gunicorn cerca app.server) ──────────────────
server = app.server

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 8050))
    app.run(host="0.0.0.0", port=port, debug=False)
